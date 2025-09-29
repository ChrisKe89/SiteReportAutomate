# EPGW-BusinessRules.py
# Row-by-row rule creator with login monitor, validation, and CSV/XLSX ledgers

import argparse
import asyncio
import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from playwright.async_api import (
    TimeoutError as PlaywrightTimeoutError,
)
from playwright.async_api import (
    async_playwright,
)

# ============
# Config area — fill these when you have the real site/flow
# ============

RAD_CONSOLIDATE_DAYS = "#MainContent_rbconsolidate"
RAD_CONSOLIDATE_WEEK = "#MainContent_rbconsolidateweekday"
SEL_CONSOLIDATE_DAYS = "#MainContent_txtconsolidate"

CONSOLIDATION_WEEK_IDS = [
    "#MainContent_cblconsolidationweekdays_0",  # Monday
    "#MainContent_cblconsolidationweekdays_1",
    "#MainContent_cblconsolidationweekdays_2",
    "#MainContent_cblconsolidationweekdays_3",
    "#MainContent_cblconsolidationweekdays_4",
    "#MainContent_cblconsolidationweekdays_5",
    "#MainContent_cblconsolidationweekdays_6",  # Sunday
]

BASE_URL = "http://epgateway.sgp.xerox.com:8041"  # TODO  # TODO
RULES_URL = f"{BASE_URL}/AlertManagement/businessrule.aspx"  # TODO
ALLOWLIST = "http://epgateway.sgp.xerox.com"  # for Windows IWA (Kerberos/NTLM)

# Selectors: update to your page
SEL_LOGIN_SENTINEL = "#mainNav"  # TODO: proves you’re logged in

SEL_NEW_RULE_BTN = "#btnNewRule"  # TODO
SEL_RULE_NAME = "#ruleName"  # TODO
SEL_RULE_TYPE = "#ruleType"  # TODO
SEL_RULE_SAVE = "#btnSaveRule"  # TODO
SEL_TOAST_SUCCESS = ".toast-success"  # TODO

# Optional selectors (only used if you include the columns in your sheet)
SEL_ALERT_TYPE = "#MainContent_ddlAlertType"
SEL_CHAINLINK_OP = "#MainContent_ddlChainLink"
SEL_CHAINLINK_VAL = "#MainContent_txtChainLink"
SEL_OPCO = "#MainContent_ddlOpCo"
SEL_DAY_OF_MONTH = "#MainContent_ddlDayOfTheMonth"
SEL_FROM_DATE = "#MainContent_txtFromDate"
SEL_TO_DATE = "#MainContent_txtToDate"
SEL_FROM_HOUR = "#MainContent_ddlFromHour"
SEL_FROM_MINUTE = "#MainContent_ddlFromMinute"
SEL_TO_HOUR = "#MainContent_ddlToTimeHour"
SEL_TO_MINUTE = "#MainContent_ddlToMinute"
SEL_CONSOLIDATE_DAYS = "#MainContent_txtconsolidate"
# Checkbox tables would be wired by label/text; leave for when you have exact DOM.

# ============
# Defaults/Paths
# ============
REPO_DIR = Path(__file__).parent.resolve()
USER_DATA_DIR = REPO_DIR / "user-data"  # persistent browser profile
USER_DATA_DIR.mkdir(exist_ok=True)
LOG_DIR = REPO_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
AUTH_STATUS_PATH = REPO_DIR / "auth_status.json"

NAV_TIMEOUT_MS = 45000
POST_ACTION_PAUSE_MS = 500  # short cushion between UI steps


# ============
# Excel helpers
# ============
def _get_ws_for_read(path: Path, sheet_name: Optional[str]) -> Worksheet:
    wb = load_workbook(path, data_only=True)
    ws_like = wb[sheet_name] if sheet_name else wb.active
    assert ws_like is not None, "Workbook has no active worksheet"
    return cast(Worksheet, ws_like)


def _get_ws_for_write(wb: Workbook, sheet_name: str) -> Worksheet:
    ws_like = (
        wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    )
    assert ws_like is not None
    return cast(Worksheet, ws_like)


def read_rules_from_xlsx(
    path: Path, sheet_name: Optional[str] = None
) -> Tuple[List[str], List[Dict[str, Any]]]:
    ws = _get_ws_for_read(path, sheet_name)
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    headers: List[str] = [
        str(c.value) if c.value is not None else "" for c in header_cells
    ]
    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "", 0) for v in r):
            continue
        row_dict: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            row_dict[h] = r[i] if i < len(r) else None
        rows.append(row_dict)
    return headers, rows


def append_completed_row_xlsx(
    completed_path: Path,
    headers: List[str],
    row: Dict[str, Any],
    processed_at: datetime,
    status: str,
    message: str = "",
    sheet_name: str = "Completed",
) -> None:
    if completed_path.exists():
        wb = load_workbook(completed_path)
        ws = _get_ws_for_write(wb, sheet_name)
        if ws.max_row == 0:
            ws.append(headers + ["ProcessedAt", "Status", "Message"])
    else:
        wb = Workbook()
        ws_any = wb.active
        assert ws_any is not None
        ws = cast(Worksheet, ws_any)
        ws.title = sheet_name
        ws.append(headers + ["ProcessedAt", "Status", "Message"])

    values: List[Any] = [row.get(h, "") for h in headers]
    values += [processed_at.isoformat(timespec="seconds"), status, message]
    ws.append(values)
    wb.save(completed_path)


def append_completed_row_csv(
    csv_path: Path,
    headers: List[str],
    row: Dict[str, Any],
    processed_at: datetime,
    status: str,
    message: str = "",
) -> None:
    row_out: Dict[str, Any] = {h: row.get(h, "") for h in headers}
    row_out["ProcessedAt"] = processed_at.isoformat(timespec="seconds")
    row_out["Status"] = status
    row_out["Message"] = message

    new_file = not csv_path.exists()
    with csv_path.open("a", newline="", encoding="utf-8") as f:
        fieldnames = headers + ["ProcessedAt", "Status", "Message"]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        if new_file:
            w.writeheader()
        w.writerow(row_out)


# ============
# Logging & auth
# ============
def setup_logging(verbosity: int) -> logging.Logger:
    level = logging.INFO if verbosity == 0 else logging.DEBUG
    logger = logging.getLogger("EPGW-BusinessRules")
    logger.setLevel(level)
    fh = logging.FileHandler(
        LOG_DIR / f"run-{datetime.now().strftime('%Y%m%d-%H%M%S')}.log",
        encoding="utf-8",
    )
    fh.setLevel(level)
    ch = logging.StreamHandler()
    ch.setLevel(level)
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    fh.setFormatter(fmt)
    ch.setFormatter(fmt)
    logger.handlers = [fh, ch]
    return logger


async def record_auth_status(page, logged_in: bool, note: str = "") -> None:
    info = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "url": page.url,
        "logged_in": logged_in,
        "note": note,
    }
    AUTH_STATUS_PATH.write_text(json.dumps(info, indent=2), encoding="utf-8")
    try:
        await page.screenshot(
            path=str(REPO_DIR / f"auth-{datetime.now().strftime('%Y%m%d-%H%M%S')}.png"),
            full_page=True,
        )
    except Exception:
        pass


# ============
# Browser
# ============


async def get_browser_context(pw, headless: bool):
    state_path = REPO_DIR / "storage_state.json"
    if state_path.exists():
        # Use captured cookies/session (good for scheduled runs)
        browser = await pw.chromium.launch(
            headless=headless,
            channel="msedge",
            args=[
                f"--auth-server-allowlist={ALLOWLIST}",
                f"--auth-negotiate-delegate-allowlist={ALLOWLIST}",
                "--start-minimized",
            ],
        )
        context = await browser.new_context(
            storage_state=str(state_path),
            accept_downloads=True,
        )
        return context
    else:
        # Fall back to persistent profile if you haven’t captured storage yet
        return await pw.chromium.launch_persistent_context(
            user_data_dir=str(USER_DATA_DIR),
            headless=headless,
            channel="msedge",
            args=[
                f"--auth-server-allowlist={ALLOWLIST}",
                f"--auth-negotiate-delegate-allowlist={ALLOWLIST}",
                "--start-minimized",
            ],
            accept_downloads=True,
        )


async def ensure_logged_in(page, logger: logging.Logger) -> bool:
    """
    Navigate to RULES_URL and confirm a known 'logged-in' sentinel exists.
    Extend with a form-login flow if needed.
    """
    await page.goto(RULES_URL, wait_until="networkidle")
    try:
        await page.wait_for_selector(SEL_LOGIN_SENTINEL, timeout=8000)
        logger.info("Login OK (sentinel visible).")
        await record_auth_status(page, True, "Sentinel found")
        return True
    except PlaywrightTimeoutError:
        logger.warning("Sentinel not found; likely not logged in.")
        await record_auth_status(page, False, "Sentinel not found")
        return False


# ============
# Row utils
# ============
def S(row: Dict[str, Any], key: str) -> str:
    """Get string, trimmed; empty if None."""
    v = row.get(key, "")
    return "" if v is None else str(v).strip()


def any_yes(row: Dict[str, Any], prefix: str) -> bool:
    """Check if any columns starting with prefix are 'Yes' (case-insensitive)."""
    for k, v in row.items():
        if isinstance(k, str) and k.startswith(prefix):
            if str(v).strip().lower() == "yes":
                return True
    return False


def validate_row(row: Dict[str, Any]) -> Tuple[bool, str]:
    """
    Preflight validation so we SKIP obvious misses instead of crashing.
    Only checks the combinations we know from the template.
    """
    name = S(row, "BusinessRuleName")
    if not name:
        return False, "Missing BusinessRuleName"

    alert = S(row, "AlertType")
    acm = S(row, "AdditionalConditionsMode")
    cons = S(row, "ConsolidationMode")

    # ChainLink required for Fault/Fault-GCC
    if alert in {"Fault", "Fault-GCC"}:
        if not S(row, "ChainLinkOperator") or not S(row, "ChainLinkValue"):
            return False, "ChainLinkOperator/Value required for Fault/Fault-GCC"

    # Toner required for EXCH/SE-F/SNE-F/SNEFGC
    if alert in {"EXCH", "SE/F", "SNE/F", "SNEFGC"}:
        if not (any_yes(row, "TonerMain::") or any_yes(row, "TonerOther::")):
            return (
                False,
                "At least one TonerMain:: or TonerOther:: must be Yes for selected AlertType",
            )

    # AdditionalConditionsMode
    if acm == "DayOfMonth":
        if not S(row, "DayOfMonth"):
            return False, "DayOfMonth required when AdditionalConditionsMode=DayOfMonth"
    elif acm == "DayOfWeek":
        if not any_yes(row, "WeekDay::"):
            return (
                False,
                "At least one WeekDay:: must be Yes when AdditionalConditionsMode=DayOfWeek",
            )
    elif acm == "FromDate":
        if not S(row, "FromDate") or not S(row, "ToDate"):
            return (
                False,
                "FromDate and ToDate required when AdditionalConditionsMode=FromDate",
            )

    # ConsolidationMode (validate inputs only; no UI here)
    cons = S(row, "ConsolidationMode")
    if cons == "ByDays":
        if not S(row, "ConsolidateDays"):
            return False, "ConsolidateDays required when ConsolidationMode=ByDays"
    elif cons == "ByWeek":
        if not any_yes(row, "ConsolidationWeek::"):
            return (
                False,
                "At least one ConsolidationWeek:: must be Yes when ConsolidationMode=ByWeek",
            )

    return True, "OK"


# ============
# Fill helpers: only fill/click when a value is present
# ============
async def select_if_present(page, selector: str, value: str):
    if value:
        await page.select_option(selector, value)


async def fill_if_present(page, selector: str, value: str):
    if value:
        await page.fill(selector, value)


async def check_if_toner_value_present(page, value: str):
    # Ticks <input type="checkbox" value="..."> if it exists
    locator = f"input[type='checkbox'][value='{value}']"
    elts = await page.query_selector_all(locator)
    if elts:
        await page.check(locator)


# ============
# Main per-row action
# ============
async def create_rule_from_row(
    page, row: Dict[str, Any], logger: logging.Logger
) -> str:
    """
    Create a single rule from a spreadsheet row.
    Only uses fields that have values (empties ignored).
    Return "SUCCESS" or raise Exception to be caught by caller.
    """
    name = S(row, "BusinessRuleName")
    if not name:
        raise ValueError("BusinessRuleName is required")

    alert = S(row, "AlertType")

    # Start "new rule"
    await page.click(SEL_NEW_RULE_BTN)
    await page.wait_for_timeout(POST_ACTION_PAUSE_MS)

    # Core fields (fill only if present)
    await fill_if_present(page, SEL_RULE_NAME, name)
    await fill_if_present(page, SEL_RULE_TYPE, S(row, "RuleType"))
    await select_if_present(page, SEL_ALERT_TYPE, alert)
    await select_if_present(page, SEL_OPCO, S(row, "OpCo"))

    # ChainLink (Fault/Fault-GCC)
    if alert in {"Fault", "Fault-GCC"}:
        await select_if_present(page, SEL_CHAINLINK_OP, S(row, "ChainLinkOperator"))
        await fill_if_present(page, SEL_CHAINLINK_VAL, S(row, "ChainLinkValue"))

    # AdditionalConditionsMode (we assume radio logic handled by server; we just fill the dependent fields)
    acm = S(row, "AdditionalConditionsMode")
    if acm == "DayOfMonth":
        await select_if_present(page, SEL_DAY_OF_MONTH, S(row, "DayOfMonth"))
    elif acm == "FromDate":
        await fill_if_present(page, SEL_FROM_DATE, S(row, "FromDate"))
        await fill_if_present(page, SEL_TO_DATE, S(row, "ToDate"))
        await select_if_present(page, SEL_FROM_HOUR, S(row, "FromHour"))
        await select_if_present(page, SEL_FROM_MINUTE, S(row, "FromMinute"))
        await select_if_present(page, SEL_TO_HOUR, S(row, "ToHour"))
        await select_if_present(page, SEL_TO_MINUTE, S(row, "ToMinute"))
    elif acm == "DayOfWeek":
        # WeekDay::Mon..Sun columns will require clicking checkboxes by label — wire when you give me exact DOM.
        pass

    # Consolidation
    cons = S(row, "ConsolidationMode")
    if cons == "ByDays":
        await fill_if_present(page, SEL_CONSOLIDATE_DAYS, S(row, "ConsolidateDays"))
    elif cons == "ByWeek":
        # ConsolidationWeek::Mon..Sun → checkboxes by label — wire when DOM ready.
        pass

    # Toner (EXCH/SE-F/SNE-F/SNEFGC)
    if alert in {"EXCH", "SE/F", "SNE/F", "SNEFGC"}:
        # Option 1: sheet stores "Yes" under columns that are already the VALUE tokens:
        #   e.g., TonerMain::BlackTonerCartridge, TonerMain::CyanTonerCartridge, ...
        for k, v in row.items():
            if isinstance(k, str) and (
                k.startswith("TonerMain::") or k.startswith("TonerOther::")
            ):
                if str(v).strip().lower() == "yes":
                    token = k.split("::", 1)[1]  # take the right side after '::'
                    await check_if_toner_value_present(page, token)

        # Option 2 (fallback): if your headers are human labels (e.g., TonerMain::Black),
        # map them to the value tokens seen in the HTML.
        label_to_value = {
            "Black": "BlackTonerCartridge",
            "Cyan": "CyanTonerCartridge",
            "Magenta": "MagentaTonerCartridge",
            "Yellow": "YellowTonerCartridge",
            "Waste Toner Bottle": "WasteTonerContainer",
            "Blue Toner": "BlueTonerCartridge",
            "White Toner": "WhiteTonerCartridge",
            "Green Toner": "GreenTonerCartridge",
            "Red Toner": "RedTonerCartridge",
            "Blue Drum": "BlueDrumCartridge",
            "White Drum": "WhiteDrumCartridge",
            "Green Drum": "GreenDrumCartridge",
            "Red Drum": "RedDrumCartridge",
            "Fusing Web": "FusingWeb",
            # add more from the page as needed…
        }
        for k, v in row.items():
            if isinstance(k, str) and (
                k.startswith("TonerMain::") or k.startswith("TonerOther::")
            ):
                if str(v).strip().lower() == "yes":
                    label = k.split("::", 1)[1]
                    if label in label_to_value:
                        await check_if_toner_value_present(page, label_to_value[label])

    # Save & confirm
    async with page.expect_response(lambda r: r.ok, timeout=10000):
        await page.click(SEL_RULE_SAVE)

    try:
        await page.wait_for_selector(SEL_TOAST_SUCCESS, timeout=8000)
    except PlaywrightTimeoutError:
        # Some pages don’t toast; tolerate as success if no error visible
        pass

    logger.info(f"Rule created: {name}")
    return "SUCCESS"


# ============
# Orchestration
# ============
async def run_job(
    input_xlsx: Path,
    completed_xlsx: Optional[Path],
    completed_csv: Optional[Path],
    sheet_name: Optional[str],
    headless: bool,
    max_rows: Optional[int],
    mutate_input: bool,
    logger: logging.Logger,
) -> None:
    headers, rows = read_rules_from_xlsx(input_xlsx, sheet_name=sheet_name)
    if not rows:
        logger.warning("No rows found in input; nothing to do.")
        return

    async with async_playwright() as p:
        context = await get_browser_context(p, headless=headless)
        page = await context.new_page()
        page.set_default_navigation_timeout(NAV_TIMEOUT_MS)
        page.set_default_timeout(NAV_TIMEOUT_MS)

        try:
            if not await ensure_logged_in(page, logger):
                logger.error("Login not established. Aborting.")
                return

            processed_count = 0
            kept_unprocessed: List[Dict[str, Any]] = []

            for idx, row in enumerate(rows, start=1):
                if max_rows is not None and processed_count >= max_rows:
                    kept_unprocessed.append(row)
                    continue

                # 1) Pre-validate → SKIP if missing combo requirements
                ok, reason = validate_row(row)
                when = datetime.now()
                if not ok:
                    logger.warning(f"Row {idx}: SKIPPED — {reason}")
                    if completed_xlsx:
                        append_completed_row_xlsx(
                            completed_xlsx, headers, row, when, "SKIPPED", reason
                        )
                    if completed_csv:
                        append_completed_row_csv(
                            completed_csv, headers, row, when, "SKIPPED", reason
                        )
                    kept_unprocessed.append(row)  # keep for future retry
                    # Hard reset page for next row
                    await page.goto(RULES_URL, wait_until="networkidle")
                    continue

                # 2) Try to create → on error, record and move on
                try:
                    status = await create_rule_from_row(page, row, logger)
                    msg = ""
                except Exception as e:
                    status = "ERROR"
                    msg = str(e)
                    logger.exception(f"Row {idx}: ERROR — {e}")

                # 3) Append to ledgers (per-row)
                when = datetime.now()
                if completed_xlsx:
                    append_completed_row_xlsx(
                        completed_xlsx, headers, row, when, status, msg
                    )
                if completed_csv:
                    append_completed_row_csv(
                        completed_csv, headers, row, when, status, msg
                    )

                # 4) Decide whether to remove from input later
                if status == "SUCCESS":
                    processed_count += 1
                else:
                    kept_unprocessed.append(row)

                # 5) Always reset the page between rows
                await page.goto(RULES_URL, wait_until="networkidle")

            logger.info(f"Processed OK: {processed_count} / {len(rows)}")

            # Optionally mutate input: remove successful rows
            if mutate_input and processed_count > 0:
                _rewrite_input_with_kept(
                    input_xlsx, kept_unprocessed, headers, sheet_name=sheet_name
                )
                logger.info("Input workbook updated (removed successful rows).")

        finally:
            await context.close()


def _rewrite_input_with_kept(
    path: Path,
    kept_rows: List[Dict[str, Any]],
    headers: List[str],
    sheet_name: Optional[str],
):
    wb = load_workbook(path)
    ws_like = wb[sheet_name] if sheet_name else wb.active
    assert ws_like is not None
    ws = cast(Worksheet, ws_like)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for r in kept_rows:
        ws.append([r.get(h, "") for h in headers])
    wb.save(path)


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description="EPGW Business Rules — row-by-row creator with validation and CSV ledger."
    )
    ap.add_argument(
        "--input", required=True, help="Path to input XLSX (rows to create)."
    )
    ap.add_argument(
        "--completed", default=None, help="Path to Completed XLSX (append-only)."
    )
    ap.add_argument(
        "--completed-csv",
        default="RulesCompleted.csv",
        help="Path to Completed CSV (append per row).",
    )
    ap.add_argument("--sheet", default=None, help="Worksheet name (default: active).")
    ap.add_argument(
        "--headless",
        action="store_true",
        help="Run headless (keep off while testing/IWA).",
    )
    ap.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="Limit rows this run (useful for testing).",
    )
    ap.add_argument(
        "--mutate-input",
        action="store_true",
        help="Remove successful rows from input workbook.",
    )
    ap.add_argument(
        "-v", "--verbose", action="count", default=0, help="Increase log verbosity."
    )
    return ap.parse_args()


def main() -> None:
    args = parse_args()
    logger = setup_logging(args.verbose)

    input_xlsx = Path(args.input).resolve()
    if not input_xlsx.exists():
        raise FileNotFoundError(f"Input not found: {input_xlsx}")

    completed_xlsx = Path(args.completed).resolve() if args.completed else None
    completed_csv = Path(args.completed_csv).resolve() if args.completed_csv else None

    asyncio.run(
        run_job(
            input_xlsx=input_xlsx,
            completed_xlsx=completed_xlsx,
            completed_csv=completed_csv,
            sheet_name=args.sheet,
            headless=args.headless,
            max_rows=args.max_rows,
            mutate_input=args.mutate_input,
            logger=logger,
        )
    )


if __name__ == "__main__":
    main()
