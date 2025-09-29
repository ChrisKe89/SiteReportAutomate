import asyncio
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List, Tuple

from bs4 import BeautifulSoup
from openpyxl import Workbook
from playwright.async_api import (
    async_playwright,
    TimeoutError as PlaywrightTimeoutError,
)

# --- Site config ---
BASE_URL = "https://sgpaphq-epbbcs3.dc01.fujixerox.net"
REPORT_URL = f"{BASE_URL}/firmware/DeviceList.aspx"

# --- Paths ---
REPO_DIR = Path(__file__).parent.resolve()
DOWNLOAD_DIR = REPO_DIR / "downloads"
DOWNLOAD_DIR.mkdir(exist_ok=True)
USER_DATA_DIR = REPO_DIR / "user-data"
USER_DATA_DIR.mkdir(exist_ok=True)

# --- Selectors ---
DDL_OPCO = "#MainContent_ddlOpCoCode"
BTN_SEARCH = "#MainContent_btnSearch"
BTN_EXPORT = "#MainContent_btnExport"

# --- Options ---
HEADLESS = False
ALLOWLIST = "*.fujixerox.net"
NAV_TIMEOUT_MS = 45000
AFTER_SEARCH_WAIT_MS = 3000

# =========================
# HTML-in-.XLS cleaner
# =========================
_XML_DECLARATION_RE = re.compile(r"<\?xml[^>]*\?>", re.IGNORECASE)
_XML_BLOCK_RE = re.compile(r"<xml[^>]*>.*?</xml>", re.IGNORECASE | re.DOTALL)


def _strip_xml_fragments(text: str) -> str:
    if not text:
        return ""
    cleaned = _XML_DECLARATION_RE.sub("", text)
    cleaned = _XML_BLOCK_RE.sub("", cleaned)
    return cleaned


def _clean_cell_text(text: str | None) -> str:
    if not text:
        return ""
    cleaned = text.replace("\xa0", " ")
    cleaned = _strip_xml_fragments(cleaned)
    return cleaned.strip()


def _extract_table(html: str) -> Tuple[List[str], List[List[str]]]:
    """Extract header + rows from the first meaningful table."""
    soup = BeautifulSoup(_strip_xml_fragments(html), "html.parser")

    # Remove residual <xml> tags if any
    for xml_tag in soup.find_all("xml"):
        xml_tag.decompose()

    tables = soup.find_all("table")
    if not tables:
        raise ValueError("No <table> elements found in the uploaded file.")

    preferred = soup.find("table", id="MainContent_gvDeviceList")
    table = (
        preferred
        if preferred
        else max(
            tables,
            key=lambda t: len(t.find("tr").find_all(["th", "td"]))
            if t.find("tr")
            else 0,
        )
    )

    trs = table.find_all("tr")
    if not trs:
        raise ValueError("The table contains no rows.")

    header_cells: List = []
    header_index: int = 0
    for idx, tr in enumerate(trs):
        ths = tr.find_all("th")
        if ths:
            header_cells = ths
            header_index = idx
            break

    if not header_cells:
        first = trs[0]
        header_cells = first.find_all(["th", "td"])
        header_index = 0

    headers = [
        _clean_cell_text(cell.get_text(separator=" ", strip=True))
        for cell in header_cells
    ]
    if not headers:
        raise ValueError("Could not determine table headers.")

    data_rows: List[List[str]] = []
    for tr in trs[header_index + 1 :]:
        tds = tr.find_all("td")
        if not tds:
            continue
        values = [
            _clean_cell_text(td.get_text(separator=" ", strip=True)) for td in tds
        ]
        if len(values) < len(headers):
            values += [""] * (len(headers) - len(values))
        elif len(values) > len(headers):
            values = values[: len(headers)]
        data_rows.append(values)

    return headers, data_rows


def clean_html_xls_to_xlsx(file_bytes: bytes, sheet_name: str = "Data") -> bytes:
    """Convert HTML-in-.xls to clean .xlsx and return XLSX bytes."""
    try:
        html = file_bytes.decode("utf-8", errors="replace")
    except Exception:
        html = file_bytes.decode("latin-1", errors="replace")

    headers, rows = _extract_table(html)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name[:31]

    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c_idx, value=str(h) if h else "")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=str(val) if val else "")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# Playwright automation
# =========================
async def run_once():
    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(USER_DATA_DIR),
            headless=HEADLESS,
            channel="msedge",  # comment if Edge channel not installed
            args=[
                f"--auth-server-allowlist={ALLOWLIST}",
                f"--auth-negotiate-delegate-allowlist={ALLOWLIST}",
                "--start-minimized",
            ],
            accept_downloads=True,
        )

        try:
            page = await context.new_page()
            page.set_default_navigation_timeout(NAV_TIMEOUT_MS)
            page.set_default_timeout(NAV_TIMEOUT_MS)

            # Navigate (IWA auto-auth)
            await page.goto(REPORT_URL, wait_until="networkidle")

            # Select FBAU
            await page.select_option(DDL_OPCO, "FXAU")

            # Search and wait
            await page.click(BTN_SEARCH)
            try:
                await page.wait_for_load_state("networkidle")
            except PlaywrightTimeoutError:
                pass
            await page.wait_for_timeout(AFTER_SEARCH_WAIT_MS)

            # Export and wait for download
            async with page.expect_download() as download_info:
                await page.click(BTN_EXPORT)
            download = await download_info.value

            # Save raw file
            stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            raw_path = DOWNLOAD_DIR / f"{stamp}-{download.suggested_filename}"
            await download.save_as(raw_path)
            print(f"[OK] Saved raw: {raw_path}")

            # Clean and save .xlsx
            try:
                xlsx_bytes = clean_html_xls_to_xlsx(
                    raw_path.read_bytes(), sheet_name="FBAU_DeviceList"
                )
                cleaned_path = raw_path.with_suffix(".xlsx")
                cleaned_path.write_bytes(xlsx_bytes)
                print(f"[OK] Saved cleaned XLSX: {cleaned_path}")
            except Exception as e:
                print(f"[WARN] Cleaning failed; raw file kept. Error: {e}")

        finally:
            await context.close()


if __name__ == "__main__":
    asyncio.run(run_once())
