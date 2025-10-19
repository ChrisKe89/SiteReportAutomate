"""Automate firmware scheduling via Playwright.

This script reads device rows from an Excel workbook, searches the
SingleRequest firmware page, and schedules upgrades when eligible.
"""
from __future__ import annotations

import asyncio
import json
import os
import random
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, Optional

from dotenv import load_dotenv  # type: ignore[import]
from openpyxl import Workbook, load_workbook  # type: ignore[import]
from playwright.async_api import (  # type: ignore[import]
    Browser,
    BrowserContext,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    async_playwright,
)

load_dotenv()

DEFAULT_URL = "https://sgpaphq-epbbcs3.dc01.fujixerox.net/firmware/SingleRequest.aspx"
INPUT_XLSX = Path(os.getenv("FIRMWARE_INPUT_XLSX", "downloads/VIC.xlsx"))
LOG_XLSX = Path(os.getenv("FIRMWARE_LOG_XLSX", "downloads/FirmwareLog.xlsx"))
BROWSER_CHANNEL = os.getenv("FIRMWARE_BROWSER_CHANNEL", "msedge")
ERRORS_JSON = Path(os.getenv("FIRMWARE_ERRORS_JSON", "errors.json")).expanduser()

STORAGE_STATE_ENV = "FIRMWARE_STORAGE_STATE"
_DEFAULT_STORAGE_STATE = "storage_state.json"

OPCO_VALUE = "FXAU"  # FBAU label

# Map uppercase state abbreviations to timezone dropdown values
TIMEZONE_BY_STATE = {
    "NT": "+09:30",
    "SA": "+10:30",
    "ACT": "+11:00",
    "VIC": "+11:00",
    "NSW": "+11:00",
    "QLD": "+10:00",
    "TAS": "+11:00",
}


@dataclass
class DeviceRow:
    serial_number: str
    product_code: str
    state: str
    opco: str

    @classmethod
    def from_iterable(cls, row: Iterable[Optional[str]]) -> "DeviceRow | None":
        cells = ["" if cell is None else str(cell).strip() for cell in row]
        if len(cells) < 4:
            return None
        serial_number, product_code, opco, state = cells[:4]
        if not serial_number or not product_code:
            return None
        return cls(
            serial_number=serial_number,
            product_code=product_code,
            opco=opco or "",
            state=state.upper() if state else "",
        )


def resolve_storage_state() -> Path:
    """Locate the persisted Playwright storage state required for login reuse."""

    env_value = os.getenv(STORAGE_STATE_ENV)
    if env_value:
        candidate = Path(env_value).expanduser()
        if candidate.is_file():
            return candidate
        raise FileNotFoundError(
            f"{STORAGE_STATE_ENV} was set to '{candidate}', but the file does not exist."
        )

    default_candidate = Path(_DEFAULT_STORAGE_STATE)
    if default_candidate.is_file():
        return default_candidate

    raise FileNotFoundError(
        "No Playwright storage state was found. Run login_capture_epgw.py to save a session before scheduling firmware."
    )


async def ensure_option_selected(page: Page, selector: str, value: str, *, timeout: float = 10_000) -> None:
    await page.wait_for_selector(selector, state="visible", timeout=timeout)
    await page.select_option(selector, value=value)


async def fill_input(page: Page, selector: str, value: str) -> None:
    locator = page.locator(selector)
    await locator.wait_for(state="visible")
    await locator.fill("")
    await locator.type(value)


def load_rows(path: Path) -> list[DeviceRow]:
    if not path.exists():
        raise FileNotFoundError(f"Input workbook not found: {path}")
    workbook = load_workbook(path, read_only=True)
    sheet = workbook.active
    rows: list[DeviceRow] = []
    # Skip header row by using min_row=2
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        item = DeviceRow.from_iterable(raw)
        if item:
            rows.append(item)
    workbook.close()
    return rows


def pick_random_schedule_date() -> datetime:
    """Return a random date between tomorrow and six days from today (inclusive)."""
    start = datetime.now().date() + timedelta(days=1)
    end = datetime.now().date() + timedelta(days=6)
    span = (end - start).days
    offset = random.randint(0, span)
    return datetime.combine(start + timedelta(days=offset), datetime.min.time())


def pick_time_option(options: list[tuple[str, str]]) -> tuple[str, str] | None:
    """Pick a random time option in allowed windows from (value, label) pairs."""
    allowed: list[tuple[str, str]] = []
    for value, label in options:
        if not value or value.lower() == "select":
            continue
        match = re.search(r"(\d{1,2})(?::(\d{2}))?\s*([ap]m)?", label, re.IGNORECASE)
        if not match:
            match = re.search(r"(\d{1,2}):(\d{2})", value)
        if not match:
            continue
        hour = int(match.group(1))
        minute = int(match.group(2)) if match.lastindex and match.lastindex >= 2 and match.group(2) else 0
        meridiem = match.group(3).lower() if match.lastindex and match.lastindex >= 3 and match.group(3) else ""
        if meridiem:
            if meridiem == "pm" and hour != 12:
                hour += 12
            if meridiem == "am" and hour == 12:
                hour = 0
        hour = hour % 24
        if 0 <= hour <= 7 or 18 <= hour <= 23:
            allowed.append((value, label))
    if not allowed:
        return None
    return random.choice(allowed)


def append_log(row: DeviceRow, status: str, message: str, *, scheduled_date: str = "", scheduled_time: str = "", timezone: str = "") -> None:
    LOG_XLSX.parent.mkdir(parents=True, exist_ok=True)
    if LOG_XLSX.exists():
        workbook = load_workbook(LOG_XLSX)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "SerialNumber",
            "Product_Code",
            "OpcoID",
            "State",
            "Status",
            "Message",
            "ScheduledDate",
            "ScheduledTime",
            "TimeZone",
            "LoggedAt",
        ])
    sheet.append(
        [
            row.serial_number,
            row.product_code,
            row.opco,
            row.state,
            status,
            message,
            scheduled_date,
            scheduled_time,
            timezone,
            datetime.now().isoformat(timespec="seconds"),
        ]
    )
    workbook.save(LOG_XLSX)
    workbook.close()


def record_error(row: DeviceRow, status: str, message: str) -> None:
    """Append an error entry to the JSON ledger for quick troubleshooting."""

    if ERRORS_JSON.parent != Path("."):
        ERRORS_JSON.parent.mkdir(parents=True, exist_ok=True)

    entries: list[dict[str, str]] = []
    if ERRORS_JSON.exists():
        try:
            existing = json.loads(ERRORS_JSON.read_text(encoding="utf-8"))
            if isinstance(existing, list):
                entries = [entry for entry in existing if isinstance(entry, dict)]
        except json.JSONDecodeError:
            entries = []

    entries.append(
        {
            "serial_number": row.serial_number,
            "product_code": row.product_code,
            "opco": row.opco,
            "state": row.state,
            "status": status,
            "message": message,
            "logged_at": datetime.now().isoformat(timespec="seconds"),
        }
    )

    ERRORS_JSON.write_text(json.dumps(entries, indent=2), encoding="utf-8")


async def select_time(page: Page) -> tuple[str, str]:
    selectors = [
        "select#MainContent_ddlTime",
        "select#MainContent_ddlTimeSlot",
        "select[id*='ddlTimeSlot']",
        "select[id*='ddlTime'][id*='Slot']",
        "select[id*='ddlTime']:not(#MainContent_ddlTimeZone)",
    ]
    for selector in selectors:
        dropdown = page.locator(selector)
        if await dropdown.count() == 0:
            continue
        options_locator = dropdown.locator("option")
        total = await options_locator.count()
        options: list[tuple[str, str]] = []
        for idx in range(total):
            option = options_locator.nth(idx)
            value = (await option.get_attribute("value")) or ""
            label = (await option.inner_text()).strip()
            options.append((value, label))
        choice = pick_time_option(options)
        if choice:
            await dropdown.select_option(choice[0])
            return choice
    raise RuntimeError("Could not determine a valid time option to select.")


async def set_schedule_date(page: Page, target: datetime) -> str:
    date_str = target.strftime("%d/%m/%Y")
    locator = page.locator("#MainContent_txtDateTime")
    await locator.wait_for(state="visible")
    await locator.evaluate(
        "(el, value) => { el.removeAttribute('readonly'); el.value = value; el.dispatchEvent(new Event('change', { bubbles: true })); }",
        date_str,
    )
    return date_str


async def select_timezone(page: Page, state: str) -> str:
    value = TIMEZONE_BY_STATE.get(state.upper())
    if not value:
        raise ValueError(f"No timezone mapping for state: {state}")
    selector = "#MainContent_ddlTimeZone"
    await page.wait_for_selector(selector, state="visible")
    await page.select_option(selector, value=value)
    return value


async def handle_row(page: Page, row: DeviceRow) -> None:
    await ensure_option_selected(page, "#MainContent_ddlOpCoID", OPCO_VALUE)
    await fill_input(page, "#MainContent_ProductCode", row.product_code)
    await fill_input(page, "#MainContent_SerialNumber", row.serial_number)
    await page.click("#MainContent_btnSearch")
    try:
        await page.wait_for_timeout(1000)
        await page.wait_for_load_state("networkidle")
    except PlaywrightTimeoutError:
        pass

    eligibility_table = page.locator("#MainContent_GridViewEligibility")
    if await eligibility_table.count() > 0 and await eligibility_table.is_visible():
        text = (await eligibility_table.inner_text()).lower()
        if "already upgraded" in text:
            append_log(row, "AlreadyUpgraded", "Device already upgraded; skipped")
            return

    device_table = page.locator("#MainContent_GridViewDevice")
    if await device_table.count() == 0 or not await device_table.is_visible():
        message = "Device table not visible after search"
        append_log(row, "NotEligible", message)
        record_error(row, "NotEligible", message)
        return

    schedule_date = pick_random_schedule_date()
    scheduled_date_str = await set_schedule_date(page, schedule_date)
    _, time_label = await select_time(page)
    timezone_value = await select_timezone(page, row.state)

    await page.click("#MainContent_submitButton")
    message_locator = page.locator("#MainContent_lblMessage, #MainContent_lblStatus")
    status_text = ""
    try:
        await message_locator.wait_for(state="visible", timeout=10_000)
        status_text = (await message_locator.inner_text()).strip()
    except PlaywrightTimeoutError:
        status_text = "No confirmation message"

    append_log(
        row,
        "Scheduled",
        status_text or "Scheduled",
        scheduled_date=scheduled_date_str,
        scheduled_time=time_label,
        timezone=timezone_value,
    )


async def run() -> None:
    rows = load_rows(INPUT_XLSX)
    if not rows:
        print("No rows to process")
        return

    try:
        storage_state_path = resolve_storage_state()
    except FileNotFoundError as exc:
        print(exc)
        return

    async with async_playwright() as p:
        browser: Browser = await p.chromium.launch(headless=False, channel=BROWSER_CHANNEL)
        context: BrowserContext = await browser.new_context(storage_state=str(storage_state_path))
        page: Page = await context.new_page()

        await page.goto(DEFAULT_URL, wait_until="domcontentloaded")
        for row in rows:
            try:
                await handle_row(page, row)
            except Exception as exc:  # noqa: BLE001
                message = str(exc)
                append_log(row, "Failed", message)
                record_error(row, "Failed", message)

        await context.storage_state(path=str(storage_state_path))
        await browser.close()


if __name__ == "__main__":
    asyncio.run(run())
