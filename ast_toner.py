"""Automate toner availability lookups and capture the results in Excel.

The script drives the Fuji Xerox parts status page via Playwright, reads input
rows from ``input.xlsx`` and persists the enriched results to ``output.xlsx``.
Robust error handling is included so transient site issues or unexpected input
structures do not crash the run.
"""

import asyncio
import logging
from typing import Iterable, List, Sequence, Tuple

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright

# Input and output file paths
INPUT_XLSX = "downloads\20251015-130256-Device List.xlsx"  # Your cleaned spreadsheet
OUTPUT_XLSX = "output.xlsx"  # Where results will be saved

# Selectors for the input fields and search button
SELECTORS = {
    "product_family": "#MainContent_txtProductFamily",
    "product_code": "#MainContent_txtProductCode",
    "serial_number": "#MainContent_txtSerialNumber",
    "search_btn": "#MainContent_btnSearch",
    "results_table": "#MainContent_gvResults",
}

PAGE_URL = "https://sgpaphq-epbbcs3.dc01.fujixerox.net/rdhc/PartStatuses.aspx"


logger = logging.getLogger(__name__)


def parse_html_table(html: str) -> Tuple[List[str], List[List[str]]]:
    """Extract header and body rows from an HTML table snippet."""

    soup = BeautifulSoup(html, "html.parser")
    table = soup.select_one("table")
    if table is None:
        raise ValueError("No <table> found in the HTML!")

    headers = [th.get_text(strip=True) for th in table.select("tr th")]
    rows: List[List[str]] = []
    for tr in table.select("tr")[1:]:
        cells = [td.get_text(strip=True) for td in tr.select("td")]
        if cells:
            rows.append(cells)

    return headers, rows


def _cell_value(row: Sequence, index: int) -> str:
    """Safely read and normalise a cell value from ``row``."""

    if index >= len(row):
        return ""
    value = row[index].value
    return str(value).strip() if value is not None else ""


def _row_has_data(values: Iterable[str]) -> bool:
    """Return True when at least one value contains user-provided data."""

    return any(value for value in values)


async def main():
    logging.basicConfig(
        format="%(asctime)s %(levelname)s %(message)s",
        level=logging.INFO,
    )

    logger.info("Loading input workbook: %s", INPUT_XLSX)
    wb = load_workbook(INPUT_XLSX)
    ws = wb.active

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Results"

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(storage_state="storage_state.json")
        page = await context.new_page()
        await page.goto(PAGE_URL)

        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            serial_number = _cell_value(row, 0)
            product_code = _cell_value(row, 1)
            product_family = _cell_value(row, 6)

            if not _row_has_data((serial_number, product_code, product_family)):
                logger.info("Skipping empty row %s", idx)
                continue

            logger.info(
                "Processing row %s: family=%s, code=%s, serial=%s",
                idx,
                product_family or "-",
                product_code or "-",
                serial_number or "-",
            )

            await page.fill(SELECTORS["product_family"], product_family)
            await page.fill(SELECTORS["product_code"], product_code)
            await page.fill(SELECTORS["serial_number"], serial_number)
            await page.click(SELECTORS["search_btn"])

            try:
                await page.wait_for_selector(SELECTORS["results_table"], timeout=10000)
                table_html = await page.inner_html(SELECTORS["results_table"])
                headers, data_rows = parse_html_table(table_html)
            except PlaywrightTimeoutError:
                logger.warning("Timed out waiting for results table on row %s", idx)
                headers, data_rows = ["Status"], [["Lookup timed out"]]
            except ValueError as exc:
                logger.warning("Failed to parse table for row %s: %s", idx, exc)
                headers, data_rows = ["Status"], [["No results table found"]]

            if idx == 2:
                out_ws.append(
                    ["Input Row", "Product Family", "Product Code", "Serial Number"]
                    + headers
                )

            if not data_rows:
                data_rows = [["No results returned"]]

            for data_row in data_rows:
                out_ws.append(
                    [idx, product_family, product_code, serial_number] + data_row
                )

        logger.info("Saving results workbook: %s", OUTPUT_XLSX)
        out_wb.save(OUTPUT_XLSX)
        await browser.close()


# To run in a notebook or interactive environment, use:
# asyncio.get_event_loop().run_until_complete(main())
# Otherwise, use:
if __name__ == "__main__":
    asyncio.run(main())
