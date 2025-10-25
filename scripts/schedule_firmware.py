"""Automate firmware scheduling via Playwright.

This script reads device rows from an Excel workbook, searches the
SingleRequest firmware page, and schedules upgrades when eligible.
"""

from __future__ import annotations

import asyncio
import csv
import json
import os
import random
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from dotenv import load_dotenv  # type: ignore[import]
from openpyxl import Workbook, load_workbook  # type: ignore[import]
from playwright.async_api import (  # type: ignore[import]
    BrowserContext,
    Error as PlaywrightError,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    async_playwright,
)

load_dotenv()

DEFAULT_URL = "https://sgpaphq-epbbcs3.dc01.fujixerox.net/firmware/SingleRequest.aspx"
def _env_path(var_name: str, default: str) -> Path:
    raw_value = os.getenv(var_name, default)
    normalised = raw_value.replace("\\", "/")
    return Path(normalised).expanduser()


INPUT_PATH = _env_path("FIRMWARE_INPUT_XLSX", "data/firmware_schedule.csv")
LOG_PATH = _env_path("FIRMWARE_LOG_XLSX", "logs/fws_log.json")
BROWSER_CHANNEL = os.getenv("FIRMWARE_BROWSER_CHANNEL", "msedge")
ERRORS_JSON = _env_path("FIRMWARE_ERRORS_JSON", "logs/fws_error_log.json")
HTTP_USERNAME = os.getenv("FIRMWARE_HTTP_USERNAME")
HTTP_PASSWORD = os.getenv("FIRMWARE_HTTP_PASSWORD")
AUTH_WARMUP_URL = os.getenv(
    "FIRMWARE_WARMUP_URL",
    "http://epgateway.sgp.xerox.com:8041/AlertManagement/businessrule.aspx",
)
ALLOWLIST = os.getenv("FIRMWARE_AUTH_ALLOWLIST", "*.fujixerox.net,*.xerox.com")
HEADLESS = os.getenv("FIRMWARE_HEADLESS", "false").lower() in {"1", "true", "yes"}
STORAGE_STATE_PATH = _env_path("FIRMWARE_STORAGE_STATE", "storage_state.json")

OPCO_VALUE = "FXAU"  # FBAU label

# Enumerate the expected timezone dropdown options and labels.
TIMEZONE_OPTIONS: list[tuple[str, str]] = [
    ("+09:30", "(UTC+09: 30) Darwin"),
    ("+10:30", "(UTC+10: 30) Adelaide"),
    ("+11:00", "(UTC+11: 00) Canberra, Melbourne, Sydney"),
    ("+10:00", "(UTC+10: 00) Brisbane"),
    ("+11:00", "(UTC+11: 00) Hobart"),
]


def _normalise_timezone_label(value: str) -> str:
    compact = re.sub(r"\s*:\s*", ":", re.sub(r"\s+", " ", value.strip()))
    return compact.upper()


EXPECTED_TIMEZONE_LABELS_NORMALISED = {
    _normalise_timezone_label(label) for _, label in TIMEZONE_OPTIONS
}
EXPECTED_TIMEZONE_VALUES = {value for value, _ in TIMEZONE_OPTIONS if value}

# Map uppercase state abbreviations to timezone dropdown values and labels.
STATE_TIMEZONE: dict[str, tuple[str, str]] = {
    "NT": ("+09:30", "(UTC+09: 30) Darwin"),
    "SA": ("+10:30", "(UTC+10: 30) Adelaide"),
    "ACT": ("+11:00", "(UTC+11: 00) Canberra, Melbourne, Sydney"),
    "VIC": ("+11:00", "(UTC+11: 00) Canberra, Melbourne, Sydney"),
    "NSW": ("+11:00", "(UTC+11: 00) Canberra, Melbourne, Sydney"),
    "QLD": ("+10:00", "(UTC+10: 00) Brisbane"),
    "TAS": ("+11:00", "(UTC+11: 00) Hobart"),
}


def _normalise_time_label(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).upper()


ALLOWED_TIME_OPTIONS: list[tuple[str, str]] = [
    ("00", "12 AM"),
    ("01", "01 AM"),
    ("02", "02 AM"),
    ("03", "03 AM"),
    ("04", "04 AM"),
    ("05", "05 AM"),
    ("06", "06 AM"),
    ("07", "07 AM"),
    ("19", "07 PM"),
    ("20", "08 PM"),
    ("21", "09 PM"),
    ("22", "10 PM"),
    ("23", "11 PM"),
]

EXPECTED_TIME_LABELS = [label for _, label in ALLOWED_TIME_OPTIONS]
EXPECTED_TIME_LABELS_NORMALISED = {
    _normalise_time_label(label) for label in EXPECTED_TIME_LABELS
}

ALLOWED_TIME_LABEL_BY_VALUE = {value: label for value, label in ALLOWED_TIME_OPTIONS}

PREFERRED_TIME_VALUE, PREFERRED_TIME_LABEL = ALLOWED_TIME_OPTIONS[0]


@dataclass
class DeviceRow:
    serial_number: str
    product_code: str
    state: str
    opco: str

    @classmethod
    def from_iterable(cls, row: Iterable[Any]) -> "DeviceRow | None":
        cells = ["" if cell is None else str(cell).strip() for cell in row]
        if len(cells) < 4:
            return None
        serial_number, product_code, opco, state = cells[:4]
        if not serial_number or not product_code:
            return None
        return cls(
            serial_number=serial_number,
            product_code=product_code,
            opco=opco or "",
            state=state.upper() if state else "",
        )


async def ensure_option_selected(
    page: Page, selector: str, value: str, *, timeout: float = 10_000
) -> None:
    await page.wait_for_selector(selector, state="visible", timeout=timeout)
    await page.select_option(selector, value=value)


async def read_status_messages(page: Page) -> str:
    """Return the latest status message shown on the page."""

    selectors = [
        "#MainContent_MessageLabel li",
        "#MainContent_MessageLabel",
        "#MainContent_lblMessage",
        "#MainContent_lblStatus",
    ]
    messages: list[str] = []

    for selector in selectors:
        locator = page.locator(selector)
        count = await locator.count()
        if count == 0:
            continue

        for idx in range(count):
            element = locator.nth(idx)
            try:
                await element.wait_for(state="visible", timeout=5_000)
            except PlaywrightTimeoutError:
                pass

            try:
                text = (await element.inner_text()).strip()
            except PlaywrightTimeoutError:
                continue
            if not text:
                continue

            normalised = re.sub(r"\s+", " ", text)
            if normalised:
                messages.append(normalised)

    unique_messages: list[str] = []
    for message in messages:
        if message not in unique_messages:
            unique_messages.append(message)

    return " | ".join(unique_messages)


async def fill_input(page: Page, selector: str, value: str) -> None:
    locator = page.locator(selector)
    await locator.wait_for(state="visible")
    await locator.click()
    await locator.fill(value)
    actual = (await locator.input_value()).strip()
    if actual != value:
        await locator.press("Control+A")
        await locator.type(value)
        actual = (await locator.input_value()).strip()
    if actual != value:
        await locator.evaluate(
            """
            (el, value) => {
              el.value = value;
              el.dispatchEvent(new Event('input', { bubbles: true }));
              el.dispatchEvent(new Event('change', { bubbles: true }));
            }
            """,
            arg=value,
        )
        actual = (await locator.input_value()).strip()
    if actual != value:
        raise RuntimeError(
            f"Unable to populate input {selector!r} with value {value!r}"
        )


def load_rows(path: Path) -> list[DeviceRow]:
    if not path.exists():
        raise FileNotFoundError(f"Input workbook not found: {path}")
    if path.suffix.lower() == ".csv":
        return _load_rows_from_csv(path)

    workbook = load_workbook(path, read_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError(f"No active worksheet found in workbook: {path}")
        rows: list[DeviceRow] = []
        for raw in sheet.iter_rows(min_row=2, values_only=True):
            item = DeviceRow.from_iterable(raw)
            if item:
                rows.append(item)
        return rows
    finally:
        workbook.close()


def _load_rows_from_csv(path: Path) -> list[DeviceRow]:
    rows: list[DeviceRow] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError(f"CSV file '{path}' does not contain a header row.")

        for raw in reader:
            normalised = {
                (key or "").strip().lower(): "" if value is None else str(value).strip()
                for key, value in raw.items()
            }

            def _value(*candidates: str) -> str:
                for candidate in candidates:
                    if candidate in normalised and normalised[candidate]:
                        return normalised[candidate]
                return ""

            serial = _value("serialnumber", "serial_number", "serial")
            product = _value("product_code", "productcode", "product")
            opco = _value("opcoid", "opco", "opco_id")
            state = _value("state")

            if not serial or not product:
                continue

            rows.append(
                DeviceRow(
                    serial_number=serial,
                    product_code=product,
                    opco=opco,
                    state=state.upper(),
                )
            )

    return rows


def pick_random_schedule_date() -> datetime:
    """Return a random date between three and six days from today (inclusive)."""
    start = datetime.now().date() + timedelta(days=3)
    end = datetime.now().date() + timedelta(days=6)
    span = (end - start).days
    offset = random.randint(0, span)
    return datetime.combine(start + timedelta(days=offset), datetime.min.time())


def pick_time_option(options: list[tuple[str, str]]) -> tuple[str, str] | None:
    """Return the first allowed option (preferring midnight) that exists."""

    cleaned_options: list[tuple[str, str]] = []
    for value, label in options:
        cleaned_value = value.strip()
        cleaned_label = label.strip()
        if not cleaned_value and not cleaned_label:
            continue
        cleaned_options.append((cleaned_value, cleaned_label))

    for allowed_value, allowed_label in ALLOWED_TIME_OPTIONS:
        allowed_normalised = _normalise_time_label(allowed_label)
        for value, label in cleaned_options:
            candidate_label = label or allowed_label
            candidate_normalised = _normalise_time_label(candidate_label)
            if candidate_normalised == allowed_normalised or (
                value and value == allowed_value
            ):
                return value or allowed_value, label or allowed_label

    return None


def append_log(
    row: DeviceRow,
    status: str,
    message: str,
    *,
    scheduled_date: str = "",
    scheduled_time: str = "",
    timezone: str = "",
) -> None:
    entry = {
        "SerialNumber": row.serial_number,
        "Product_Code": row.product_code,
        "OpcoID": row.opco,
        "State": row.state,
        "Status": status,
        "Message": message,
        "ScheduledDate": scheduled_date,
        "ScheduledTime": scheduled_time,
        "TimeZone": timezone,
        "LoggedAt": datetime.now().isoformat(timespec="seconds"),
    }

    if LOG_PATH.suffix.lower() == ".json":
        _append_log_json(entry)
    else:
        _append_log_workbook(entry)


def _append_log_json(entry: dict[str, str]) -> None:
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload: list[dict[str, str]] = []
    if LOG_PATH.exists():
        try:
            existing = json.loads(LOG_PATH.read_text(encoding="utf-8"))
            if isinstance(existing, list):
                payload = [item for item in existing if isinstance(item, dict)]
        except json.JSONDecodeError:
            payload = []
    payload.append(entry)
    LOG_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _append_log_workbook(entry: dict[str, str]) -> None:
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    if LOG_PATH.exists():
        workbook = load_workbook(LOG_PATH)
        sheet = workbook.active
        if sheet is None:
            workbook.close()
            raise ValueError("Log workbook has no active worksheet.")
    else:
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            workbook.close()
            raise ValueError("Unable to create log worksheet.")
        sheet.append(list(entry.keys()))
    sheet.append(list(entry.values()))
    workbook.save(LOG_PATH)
    workbook.close()


def record_error(row: DeviceRow, status: str, message: str) -> None:
    """Append an error entry to the JSON ledger for quick troubleshooting."""

    if ERRORS_JSON.parent != Path("."):
        ERRORS_JSON.parent.mkdir(parents=True, exist_ok=True)

    entries: list[dict[str, str]] = []
    if ERRORS_JSON.exists():
        try:
            existing = json.loads(ERRORS_JSON.read_text(encoding="utf-8"))
            if isinstance(existing, list):
                entries = [entry for entry in existing if isinstance(entry, dict)]
        except json.JSONDecodeError:
            entries = []

    entries.append(
        {
            "serial_number": row.serial_number,
            "product_code": row.product_code,
            "opco": row.opco,
            "state": row.state,
            "status": status,
            "message": message,
            "logged_at": datetime.now().isoformat(timespec="seconds"),
        }
    )

    ERRORS_JSON.write_text(json.dumps(entries, indent=2), encoding="utf-8")


def remove_row_from_input(path: Path, row: DeviceRow) -> None:
    """Remove the first matching data row from the input workbook."""

    if not path.exists():
        return

    if path.suffix.lower() == ".csv":
        _remove_row_from_csv(path, row)
        return

    workbook = load_workbook(path)
    sheet = workbook.active
    if sheet is None:
        workbook.close()
        raise ValueError("Input workbook has no active worksheet.")

    row_to_delete: int | None = None
    for idx in range(2, sheet.max_row + 1):
        serial_cell = sheet.cell(row=idx, column=1).value
        product_cell = sheet.cell(row=idx, column=2).value
        serial = "" if serial_cell is None else str(serial_cell).strip()
        product = "" if product_cell is None else str(product_cell).strip()
        if serial == row.serial_number and product == row.product_code:
            row_to_delete = idx
            break

    if row_to_delete is not None:
        sheet.delete_rows(row_to_delete)
        workbook.save(path)
    workbook.close()


def _remove_row_from_csv(path: Path, row: DeviceRow) -> None:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            return
        fieldnames = reader.fieldnames
        records = list(reader)

    def _normalise(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    target_serial = row.serial_number.strip()
    target_product = row.product_code.strip()

    index_to_remove: int | None = None
    for idx, record in enumerate(records):
        serial = _normalise(record.get("SerialNumber") or record.get("serialnumber"))
        product = _normalise(record.get("Product_Code") or record.get("product_code"))
        if serial == target_serial and product == target_product:
            index_to_remove = idx
            break

    if index_to_remove is None:
        return

    del records[index_to_remove]

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(record)


async def reset_form(page: Page) -> None:
    """Attempt to reset the firmware request form."""

    reset_selector = "#MainContent_btnReset"
    reset_locator = page.locator(reset_selector)
    if await reset_locator.count() == 0:
        return

    try:
        await reset_locator.wait_for(state="visible", timeout=2000)
        if await reset_locator.is_disabled():
            return
        await reset_locator.click()
        await page.wait_for_timeout(500)
    except PlaywrightError:
        pass


async def select_time(page: Page) -> tuple[str, str]:
    selectors = [
        "select#MainContent_ddlScheduleTime",
    ]
    for selector in selectors:
        dropdown_group = page.locator(selector)
        count = await dropdown_group.count()
        if count == 0:
            continue

        for idx in range(count):
            dropdown = dropdown_group.nth(idx)
            if not await dropdown.is_visible():
                continue
            if await dropdown.is_disabled():
                continue

            options_locator = dropdown.locator("option")
            total = await options_locator.count()
            options: list[tuple[str, str]] = []
            for option_idx in range(total):
                option = options_locator.nth(option_idx)
                value = ((await option.get_attribute("value")) or "").strip()
                label = (await option.inner_text()).strip()
                if not value and "select" in label.lower():
                    continue
                options.append((value, label))

            choice = pick_time_option(options)
            if not choice:
                continue

            value, label = choice
            cleaned_value = value.strip()
            cleaned_label = label.strip()

            target_value = cleaned_value or PREFERRED_TIME_VALUE
            target_label = cleaned_label or ALLOWED_TIME_LABEL_BY_VALUE.get(
                target_value, PREFERRED_TIME_LABEL
            )

            try:
                selected = await dropdown.select_option(value=str(target_value))
                if not selected:
                    raise PlaywrightError("no time option matched value")
            except PlaywrightError:
                try:
                    await dropdown.select_option(label=target_label)
                except PlaywrightError:
                    await dropdown.evaluate(
                        """
                        (el, payload) => {
                          el.value = payload.value;
                          const match = Array.from(el.options).find(
                            (option) =>
                              option.value === payload.value ||
                              option.text.trim() === payload.label.trim()
                          );
                          if (match) {
                            match.selected = true;
                            el.dispatchEvent(new Event('change', { bubbles: true }));
                          }
                        }
                        """,
                        arg={"value": str(target_value), "label": target_label},
                    )

            selected_value = (await dropdown.input_value()).strip() or str(target_value)

            selected_label_locator = dropdown.locator("option:checked")
            if await selected_label_locator.count() > 0:
                selected_label = (
                    await selected_label_locator.first.inner_text()
                ).strip()
            else:
                selected_label = target_label

            if not selected_label:
                selected_label = target_label

            if selected_value:
                return selected_value, selected_label
    raise RuntimeError("Could not determine a valid time option to select.")


async def set_schedule_date(page: Page, target: datetime) -> str:
    iso_date = target.strftime("%Y-%m-%d")
    target_date = target.date()
    locator = page.locator("#MainContent_txtDateTime")
    await locator.wait_for(state="visible")
    await locator.click()

    async def read_input_value() -> str:
        for _ in range(10):
            actual = (await locator.input_value()).strip()
            if actual:
                return actual
            await page.wait_for_timeout(100)
        return ""

    def normalise_date_value(value: str) -> str:
        cleaned = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                parsed = datetime.strptime(cleaned, fmt)
            except ValueError:
                continue
            return parsed.strftime("%Y-%m-%d")
        return cleaned or iso_date

    # Attempt to use a jQuery UI style date picker if present.
    datepicker = page.locator("#ui-datepicker-div")
    if await datepicker.count() > 0:
        try:
            await datepicker.wait_for(state="visible")

            async def current_month_year() -> tuple[int, int]:
                month_text = (
                    await datepicker.locator(".ui-datepicker-month").first.inner_text()
                ).strip()
                year_text = (
                    await datepicker.locator(".ui-datepicker-year").first.inner_text()
                ).strip()
                month_number = datetime.strptime(month_text, "%B").month
                return month_number, int(year_text)

            # Navigate to the month containing the target date.
            for _ in range(12):
                month_number, year_number = await current_month_year()
                if (
                    year_number == target_date.year
                    and month_number == target_date.month
                ):
                    break
                await datepicker.locator(".ui-datepicker-next").click()
                await page.wait_for_timeout(200)

            day_locator = datepicker.locator(".ui-datepicker-calendar td a").filter(
                has_text=str(target_date.day)
            )
            if await day_locator.count() > 0:
                await day_locator.first.click()
                actual = await read_input_value()
                if actual:
                    return normalise_date_value(actual)
        except Exception:  # noqa: BLE001
            pass

    # Attempt to use the ASP.NET AJAX calendar widget if present.
    ajax_calendar = page.locator("#MainContent_CalendarDateTime_container")
    if await ajax_calendar.count() > 0:
        try:
            popup = ajax_calendar.locator(".ajax__calendar_container")
            await popup.wait_for(state="visible", timeout=2_000)

            title_locator = popup.locator(".ajax__calendar_title").first

            async def current_month_year_ajax() -> tuple[int, int]:
                title_text = (await title_locator.inner_text()).strip()
                cleaned = " ".join(title_text.split())
                match = re.match(r"([A-Za-z]+),?\s+(\d{4})", cleaned)
                if not match:
                    raise ValueError(
                        f"Unexpected calendar title format: {title_text!r}"
                    )
                month_name, year_text = match.groups()
                month_number = datetime.strptime(month_name, "%B").month
                return month_number, int(year_text)

            for _ in range(24):
                month_number, year_number = await current_month_year_ajax()
                if (
                    year_number == target_date.year
                    and month_number == target_date.month
                ):
                    break
                if (year_number, month_number) < (target_date.year, target_date.month):
                    await popup.locator(
                        "#MainContent_CalendarDateTime_nextArrow"
                    ).click()
                else:
                    await popup.locator(
                        "#MainContent_CalendarDateTime_prevArrow"
                    ).click()
                await page.wait_for_timeout(200)

            day_locator = popup.locator(
                "td:not(.ajax__calendar_invalid) .ajax__calendar_day"
            ).filter(has_text=str(target_date.day))
            if await day_locator.count() > 0:
                await day_locator.first.click()
                actual = await read_input_value()
                if actual:
                    return normalise_date_value(actual)
        except PlaywrightTimeoutError:
            pass
        except Exception:  # noqa: BLE001
            pass

    # Fallback: directly set the value if date picker interaction fails.
    await locator.evaluate(
        """
        (el, value) => {
          el.removeAttribute('readonly');
          el.value = value;
          el.dispatchEvent(new Event('input', { bubbles: true }));
          el.dispatchEvent(new Event('change', { bubbles: true }));
        }
        """,
        arg=iso_date,
    )
    actual = await read_input_value()
    if actual:
        return normalise_date_value(actual)
    return iso_date


async def select_timezone(page: Page, state: str) -> tuple[str, str]:
    """Select the timezone matching the provided state and return (value, label)."""
    state_key = state.upper()
    target = STATE_TIMEZONE.get(state_key)
    if not target:
        raise ValueError(f"No timezone mapping for state: {state}")
    target_value, target_label = target
    selector = "#MainContent_ddlTimeZone"
    dropdown = page.locator(selector)
    await dropdown.wait_for(state="visible")

    options_locator = dropdown.locator("option")
    total = await options_locator.count()
    options: list[tuple[str, str]] = []
    for idx in range(total):
        option = options_locator.nth(idx)
        value = ((await option.get_attribute("value")) or "").strip()
        label = (await option.inner_text()).strip()
        if not value and "select" in label.lower():
            continue
        options.append((value, label))

    try:
        await dropdown.select_option(value=target_value)
    except PlaywrightError:
        if target_label:
            try:
                await dropdown.select_option(label=target_label)
            except PlaywrightError:
                await dropdown.evaluate(
                    """
                    (el, payload) => {
                      el.value = payload.value;
                      const match = Array.from(el.options).find((option) => option.value === payload.value);
                      if (match) {
                        match.selected = true;
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                      }
                    }
                    """,
                    arg={"value": target_value},
                )
        else:
            raise

    selected_value = (await dropdown.input_value()).strip()
    if target_value and selected_value != target_value:
        await dropdown.evaluate(
            """
            (el, value) => {
              el.value = value;
              el.dispatchEvent(new Event('change', { bubbles: true }));
            }
            """,
            arg=target_value,
        )
        selected_value = (await dropdown.input_value()).strip()

    selected_label_locator = dropdown.locator("option:checked")
    if await selected_label_locator.count() > 0:
        selected_label = (await selected_label_locator.first.inner_text()).strip()
    else:
        selected_label = target_label

    if (
        target_label
        and selected_label
        and _normalise_timezone_label(selected_label)
        != _normalise_timezone_label(target_label)
    ):
        await dropdown.evaluate(
            """
            (el, payload) => {
              const normalise = (value) =>
                value.trim().toLowerCase().replace(/\\s+/g, ' ').replace(/\\s*:\\s*/g, ':');
              const options = Array.from(el.options);
              const match = options.find(
                (option) => normalise(option.text) === payload.normalised
              );
              if (match) {
                match.selected = true;
                el.value = match.value;
                el.dispatchEvent(new Event('change', { bubbles: true }));
              }
            }
            """,
            arg={
                "label": target_label,
                "normalised": _normalise_timezone_label(target_label).lower(),
            },
        )
        selected_value = (await dropdown.input_value()).strip()
        if await selected_label_locator.count() > 0:
            selected_label = (await selected_label_locator.first.inner_text()).strip()

    return selected_value or target_value, selected_label


async def handle_row(page: Page, row: DeviceRow) -> str:
    await ensure_option_selected(page, "#MainContent_ddlOpCoID", OPCO_VALUE)
    await fill_input(page, "#MainContent_ProductCode", row.product_code)
    await fill_input(page, "#MainContent_SerialNumber", row.serial_number)
    await page.click("#MainContent_btnSearch")
    try:
        await page.wait_for_timeout(1000)
        await page.wait_for_load_state("networkidle")
    except PlaywrightTimeoutError:
        pass

    eligibility_table = page.locator("#MainContent_GridViewEligibility")
    if await eligibility_table.count() > 0 and await eligibility_table.is_visible():
        text = (await eligibility_table.inner_text()).lower()
        if "already upgraded" in text:
            append_log(row, "AlreadyUpgraded", "Device already upgraded; skipped")
            return "AlreadyUpgraded"

    device_table = page.locator("#MainContent_GridViewDevice")
    if await device_table.count() == 0 or not await device_table.is_visible():
        message = await read_status_messages(page)
        if not message:
            message = "Device table not visible after search"
        append_log(row, "NotEligible", message)
        record_error(row, "NotEligible", message)
        await reset_form(page)
        return "NotEligible"

    schedule_date = pick_random_schedule_date()
    scheduled_date_str = await set_schedule_date(page, schedule_date)
    time_value, time_label = await select_time(page)
    timezone_value, timezone_label = await select_timezone(page, row.state)

    schedule_button = page.locator("#MainContent_submitButton")
    if await schedule_button.count() == 0:
        message = "Schedule button not found; form reset triggered"
        await reset_form(page)
        append_log(row, "Failed", message)
        record_error(row, "Failed", message)
        return "Failed"

    if not await schedule_button.is_visible() or await schedule_button.is_disabled():
        message = "Schedule button unavailable; form reset triggered"
        await reset_form(page)
        append_log(row, "Failed", message)
        record_error(row, "Failed", message)
        return "Failed"

    await schedule_button.click()
    try:
        await page.wait_for_timeout(500)
    except PlaywrightTimeoutError:
        pass

    status_text = await read_status_messages(page)
    if not status_text:
        status_text = "Scheduled"

    append_log(
        row,
        "Scheduled",
        status_text,
        scheduled_date=scheduled_date_str,
        scheduled_time=time_label or time_value,
        timezone=timezone_label or timezone_value,
    )

    await reset_form(page)

    return "Scheduled"


async def run() -> None:
    rows = load_rows(INPUT_PATH)
    if not rows:
        print("No rows to process")
        return

    if not STORAGE_STATE_PATH.exists():
        raise FileNotFoundError(
            "Storage state not found. Capture a session with login_capture_epgw.py "
            f"and ensure it is available at {STORAGE_STATE_PATH}."
        )

    async with async_playwright() as p:
        launch_kwargs: dict[str, Any] = {
            "headless": HEADLESS,
            "args": [
                f"--auth-server-allowlist={ALLOWLIST}",
                f"--auth-negotiate-delegate-allowlist={ALLOWLIST}",
                "--start-minimized",
            ],
        }
        if BROWSER_CHANNEL:
            launch_kwargs["channel"] = BROWSER_CHANNEL
        browser = await p.chromium.launch(**launch_kwargs)

        context_kwargs: dict[str, Any] = {
            "storage_state": str(STORAGE_STATE_PATH),
            "accept_downloads": True,
        }
        if HTTP_USERNAME and HTTP_PASSWORD:
            context_kwargs["http_credentials"] = {
                "username": HTTP_USERNAME,
                "password": HTTP_PASSWORD,
            }

        context: BrowserContext = await browser.new_context(**context_kwargs)

        try:
            page: Page = await context.new_page()
            page.set_default_navigation_timeout(45_000)
            page.set_default_timeout(45_000)

            if AUTH_WARMUP_URL:
                try:
                    await page.goto(AUTH_WARMUP_URL, wait_until="domcontentloaded")
                except PlaywrightError as exc:
                    print(f"Warm-up navigation to {AUTH_WARMUP_URL} failed: {exc}")

            await page.goto(DEFAULT_URL, wait_until="domcontentloaded")
            for row in list(rows):
                try:
                    status = await handle_row(page, row)
                    if status == "Scheduled":
                        remove_row_from_input(INPUT_PATH, row)
                        if row in rows:
                            rows.remove(row)
                except Exception as exc:  # noqa: BLE001
                    status_text = await read_status_messages(page)
                    message = status_text or str(exc)
                    append_log(row, "Failed", message)
                    record_error(row, "Failed", message)
                    try:
                        await reset_form(page)
                    except PlaywrightError:
                        pass
        except PlaywrightError as exc:
            message = str(exc)
            if "ERR_INVALID_AUTH_CREDENTIALS" in message:
                print(
                    "Authentication failed when opening the firmware scheduling page. "
                    "Ensure Integrated Windows Authentication is available or set "
                    "FIRMWARE_HTTP_USERNAME/FIRMWARE_HTTP_PASSWORD in your .env file."
                )
            else:
                raise
        finally:
            try:
                await context.close()
            except PlaywrightError:
                pass
            await browser.close()


if __name__ == "__main__":
    asyncio.run(run())
