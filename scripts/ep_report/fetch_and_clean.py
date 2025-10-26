# fetch_and_clean.py
# Downloads the EPGW Device List report, then converts the HTML-in-.xls to a clean .xlsx.
# Patched to satisfy mypy/pylance: avoid Optional operands and cast OpenPyXL Worksheet.

import asyncio
import os
import re
import shutil
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple, cast

from bs4 import BeautifulSoup  # type: ignore[import-untyped]
from bs4.element import Tag  # type: ignore[import-untyped]
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # typed Worksheet for casts
from playwright.async_api import (
    async_playwright,
    TimeoutError as PlaywrightTimeoutError,
)
from dotenv import load_dotenv  # type: ignore[import-untyped]

load_dotenv()


def _env_path(var_name: str, default: str) -> Path:
    raw_value = os.getenv(var_name, default)
    normalised = raw_value.replace("\\", "/")
    return Path(normalised).expanduser()


# --- Site config ---
BASE_URL = os.getenv("FETCH_BASE_URL", "https://sgpaphq-epbbcs3.dc01.fujixerox.net")
REPORT_URL = os.getenv("FETCH_REPORT_URL", f"{BASE_URL}/firmware/DeviceList.aspx")

# --- Paths ---
DOWNLOAD_DIR = _env_path("FETCH_DOWNLOAD_DIR", "downloads")
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
USER_DATA_DIR = _env_path(
    "FETCH_USER_DATA_DIR", "user-data"
)  # persists browser profile (cookies, IWA trust, etc.)
USER_DATA_DIR.mkdir(parents=True, exist_ok=True)
REPORT_OUTPUT_XLSX = _env_path("REPORT_OUTPUT_XLSX", "data/EPFirmwareReport.xlsx")

# --- Selectors (from page) ---
DDL_OPCO = os.getenv("FETCH_SELECTOR_DDL_OPCO", "#MainContent_ddlOpCoCode")
BTN_SEARCH = os.getenv("FETCH_SELECTOR_BTN_SEARCH", "#MainContent_btnSearch")
BTN_EXPORT = os.getenv("FETCH_SELECTOR_BTN_EXPORT", "#MainContent_btnExport")

# --- Options ---
HEADLESS = os.getenv("FETCH_HEADLESS", "false").lower() in {"1", "true", "yes"}
ALLOWLIST = os.getenv("FETCH_AUTH_ALLOWLIST", "*.fujixerox.net")
NAV_TIMEOUT_MS = int(os.getenv("FETCH_NAV_TIMEOUT_MS", "45000"))
AFTER_SEARCH_WAIT_MS = int(os.getenv("FETCH_AFTER_SEARCH_WAIT_MS", "3000"))

# =========================
# HTML .xls -> clean .xlsx
# =========================

_XML_DECLARATION_RE = re.compile(r"<\?xml[^>]*\?>", re.IGNORECASE)
_XML_BLOCK_RE = re.compile(r"<xml[^>]*>.*?<\/xml>", re.IGNORECASE | re.DOTALL)


def _strip_xml_fragments(text: str) -> str:
    if not text:
        return ""
    cleaned = _XML_DECLARATION_RE.sub("", text)
    cleaned = _XML_BLOCK_RE.sub("", cleaned)
    return cleaned


def _clean_cell_text(text: Optional[str]) -> str:
    if not text:
        return ""
    cleaned = text.replace("\xa0", " ")
    cleaned = _strip_xml_fragments(cleaned)
    return cleaned.strip()


def _extract_table(html: str) -> Tuple[List[str], List[List[str]]]:
    """Extracts header and rows from the first significant HTML table."""
    soup = BeautifulSoup(_strip_xml_fragments(html), "html.parser")

    # Remove residual <xml> tags if any slipped through decoding
    for xml_tag in soup.find_all("xml"):
        xml_tag.decompose()

    # Ensure we treat only Tag instances as tables for type-checkers
    tables: List[Tag] = [t for t in soup.find_all("table") if isinstance(t, Tag)]
    if not tables:
        raise ValueError("No <table> elements found in the uploaded file.")

    # Prefer specific id if available
    preferred_any = soup.find("table", id="MainContent_gvDeviceList")
    table: Optional[Tag] = preferred_any if isinstance(preferred_any, Tag) else None

    if table is None:
        # Pick the table with the most columns (rough heuristic)
        def table_score(t: Tag) -> int:
            first_tr_any = t.find("tr")
            if not isinstance(first_tr_any, Tag):
                return 0
            return len(first_tr_any.find_all(["th", "td"]))

        table = max(tables, key=table_score)

    # Help type-checkers: ensure table is a Tag here
    assert isinstance(table, Tag)

    trs: List[Tag] = [tr for tr in table.find_all("tr") if isinstance(tr, Tag)]
    if not trs:
        raise ValueError("The table contains no rows.")

    # Identify header row
    header_cells: Optional[List[Tag]] = None
    header_index: Optional[int] = None
    for idx, tr in enumerate(trs):
        ths: List[Tag] = [
            in_th for in_th in tr.find_all("th") if isinstance(in_th, Tag)
        ]
        if ths:
            header_cells = ths
            header_index = idx
            break

    if header_cells is None:
        # Fallback: use the first row's cells as headers
        first = trs[0]
        header_cells = [c for c in first.find_all(["th", "td"]) if isinstance(c, Tag)]
        header_index = 0

    # At this point header_index is guaranteed set; assert for type-checkers
    assert header_index is not None, "header_index unexpectedly None"

    headers = [
        _clean_cell_text(cell.get_text(separator=" ", strip=True))
        for cell in header_cells
    ]
    if not headers:
        raise ValueError("Could not determine table headers.")

    # Extract data rows after header_index
    data_rows: List[List[str]] = []
    for tr in trs[header_index + 1 :]:
        tds: List[Tag] = [td for td in tr.find_all("td") if isinstance(td, Tag)]
        if not tds:
            continue
        values = [
            _clean_cell_text(td.get_text(separator=" ", strip=True)) for td in tds
        ]
        # Normalize number of columns to headers length
        if len(values) < len(headers):
            values += [""] * (len(headers) - len(values))
        elif len(values) > len(headers):
            values = values[: len(headers)]
        data_rows.append(values)

    return headers, data_rows


def _active_sheet(wb: Workbook) -> Worksheet:
    """Return a typed Worksheet for mypy/pylance."""
    return cast(Worksheet, wb.active)


def clean_html_xls_to_xlsx_bytes(raw_bytes: bytes, sheet_name: str = "Data") -> bytes:
    """Input: raw bytes from an HTML-in-.xls file. Output: XLSX bytes."""
    try:
        html = raw_bytes.decode("utf-8", errors="replace")
    except Exception:
        html = raw_bytes.decode("latin-1", errors="replace")

    headers, rows = _extract_table(html)

    wb = Workbook()
    ws = _active_sheet(wb)
    # Guard: sheet_name may be >31; slice and ensure string
    ws.title = str(sheet_name)[:31]

    # Write header row (as text)
    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c_idx, value=str(h) if h is not None else "")

    # Write data rows (as text)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=str(val) if val is not None else "")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ======================
# Download + Convert Run
# ======================


async def download_device_list_once() -> Path:
    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(USER_DATA_DIR),
            headless=HEADLESS,
            channel="msedge",  # comment this line if Edge channel isn't available
            args=[
                f"--auth-server-allowlist={ALLOWLIST}",
                f"--auth-negotiate-delegate-allowlist={ALLOWLIST}",
                "--start-minimized",
            ],
            accept_downloads=True,
        )

        try:
            page = await context.new_page()
            page.set_default_navigation_timeout(NAV_TIMEOUT_MS)
            page.set_default_timeout(NAV_TIMEOUT_MS)

            # 1) Go to report page (IWA should auto-auth if your Windows session has access)
            await page.goto(REPORT_URL, wait_until="networkidle")

            # 2) Set dropdown to FBAU (value="FXAU")
            await page.select_option(DDL_OPCO, "FXAU")

            # 3) Click Search and wait for results to load/settle
            await page.click(BTN_SEARCH)
            try:
                await page.wait_for_load_state("networkidle")
            except PlaywrightTimeoutError:
                pass
            await page.wait_for_timeout(AFTER_SEARCH_WAIT_MS)

            # 4) Click Export and capture the download
            async with page.expect_download() as download_info:
                await page.click(BTN_EXPORT)
            download = await download_info.value

            # 5) Save with a timestamped filename into downloads/ (always safe)
            suggested = download.suggested_filename or "report.xls"
            safe_name = Path(suggested).name  # strip any path shenanigans
            stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            out_path = DOWNLOAD_DIR / f"{stamp}-{safe_name}"
            await download.save_as(out_path)
            print(f"[OK] Saved raw report: {out_path.resolve()}")
            return out_path

        finally:
            await context.close()


async def main() -> None:
    # Step 1: download the HTML-in-.xls
    raw_path = await download_device_list_once()

    # Step 2: convert to clean .xlsx
    raw_bytes = raw_path.read_bytes()
    xlsx_bytes = clean_html_xls_to_xlsx_bytes(raw_bytes, sheet_name="DeviceList")

    xlsx_path = raw_path.with_suffix(".xlsx")  # keep same timestamped prefix
    xlsx_path.write_bytes(xlsx_bytes)
    print(f"[OK] Wrote clean XLSX: {xlsx_path.resolve()}")

    REPORT_OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    destination = REPORT_OUTPUT_XLSX
    if destination.exists():
        destination.unlink()
    moved_to = Path(shutil.move(str(xlsx_path), str(destination)))
    print(f"[OK] Moved cleaned XLSX to: {moved_to.resolve()}")


if __name__ == "__main__":
    asyncio.run(main())
