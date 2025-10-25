"""Automate RDHC AST toner lookup and export results to CSV."""

from __future__ import annotations

import asyncio
import csv
import logging
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from bs4 import BeautifulSoup  # type: ignore[import-untyped]
from dotenv import load_dotenv  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import column_index_from_string  # type: ignore[import-untyped]
from playwright.async_api import (  # type: ignore[import-untyped]
    Error as PlaywrightError,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    async_playwright,
)

from playwright_launch import launch_browser

load_dotenv()


def _env_path(var_name: str, default: str) -> Path:
    raw_value = os.getenv(var_name, default)
    normalised = raw_value.replace("\\", "/")
    return Path(normalised).expanduser()


AST_INPUT_XLSX = _env_path("AST_INPUT_XLSX", "data/EPFirmwareReport.xlsx")
AST_OUTPUT_CSV = _env_path("AST_OUTPUT_CSV", "data/AST_Toner_Levels.csv")
DEFAULT_RDHC_HTML = Path(__file__).resolve().parent.parent / "RDHC.html"
RDHC_HTML_PATH = _env_path("RDHC_HTML_PATH", str(DEFAULT_RDHC_HTML))
AST_PAGE_URL = os.getenv("AST_TONER_PAGE_URL", Path(RDHC_HTML_PATH).resolve().as_uri())
AST_STORAGE_STATE = _env_path("AST_TONER_STORAGE_STATE", "storage_state.json")
AST_BROWSER_CHANNEL = os.getenv("AST_BROWSER_CHANNEL", "")
AST_HEADLESS = os.getenv("AST_HEADLESS", "true").lower() in {"1", "true", "yes"}
PRODUCT_FAMILY_COLUMN = os.getenv(
    "PRODUCT_FAMILY_COLUMN", os.getenv("PRODUCT_FAMILY", "G")
)
PRODUCT_CODE_COLUMN = os.getenv("PRODUCT_CODE_COLUMN", os.getenv("PRODUCT_CODE", "B"))
SERIAL_COLUMN = os.getenv("SERIAL_COLUMN", os.getenv("SERIAL", "A"))

SELECTORS = {
    "product_family": "#MainContent_ddlProductFamily",
    "product_code": "#MainContent_txtProductCode",
    "serial_number": "#MainContent_txtSerialNumber",
    "submit": "#MainContent_btnSubmit",
    "result_panel": "#MainContent_UpdatePanelResult",
}


@dataclass
class InputRow:
    serial_number: str
    product_code: str
    product_family: str


def _normalise(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip()).lower()


def load_product_family_map(html_path: Path) -> dict[str, str]:
    if not html_path.exists():
        raise FileNotFoundError(
            f"RDHC HTML reference not found at {html_path}. Provide RDHC_HTML_PATH in your environment."
        )
    soup = BeautifulSoup(
        html_path.read_text(encoding="utf-8", errors="ignore"), "html.parser"
    )
    select = soup.select_one(SELECTORS["product_family"])
    if select is None:
        raise ValueError("Could not locate the product family dropdown in RDHC.html")

    mapping: dict[str, str] = {}
    for option in select.find_all("option"):
        value = (option.get("value") or "").strip()
        label = option.get_text(strip=True)
        if not value and not label:
            continue
        key_candidates = {value.lower(), label.lower()}
        for candidate in key_candidates:
            if candidate:
                mapping[candidate] = value or label
    return mapping


def _column_index(column_ref: str, *, label: str) -> int:
    try:
        return column_index_from_string(column_ref.strip())
    except ValueError as exc:  # noqa: B904
        raise ValueError(
            f"Invalid column reference '{column_ref}' for {label} column"
        ) from exc


def load_input_rows(path: Path) -> list[InputRow]:
    if not path.exists():
        raise FileNotFoundError(f"AST input workbook not found: {path}")

    logging.info("Reading AST input workbook from %s", path)

    family_col = _column_index(PRODUCT_FAMILY_COLUMN, label="Product Family")
    product_col = _column_index(PRODUCT_CODE_COLUMN, label="Product Code")
    serial_col = _column_index(SERIAL_COLUMN, label="Serial Number")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError("AST input workbook does not contain an active worksheet")

        start_col = min(serial_col, product_col, family_col)
        end_col = max(serial_col, product_col, family_col)
        serial_idx = serial_col - start_col
        product_idx = product_col - start_col
        family_idx = family_col - start_col

        rows: list[InputRow] = []
        empty_streak = 0
        for row_values in sheet.iter_rows(
            min_row=2,
            min_col=start_col,
            max_col=end_col,
            values_only=True,
        ):
            serial = (
                row_values[serial_idx] if len(row_values) > serial_idx else None
            )
            product = (
                row_values[product_idx] if len(row_values) > product_idx else None
            )
            family = (
                row_values[family_idx] if len(row_values) > family_idx else None
            )

            serial_text = "" if serial is None else str(serial).strip()
            product_text = "" if product is None else str(product).strip()
            family_text = "" if family is None else str(family).strip()

            if not serial_text and not product_text and not family_text:
                empty_streak += 1
                if rows and empty_streak >= 50:
                    break
                continue

            empty_streak = 0

            rows.append(
                InputRow(
                    serial_number=serial_text,
                    product_code=product_text,
                    product_family=family_text,
                )
            )
        return rows
    finally:
        workbook.close()


async def _fill(page: Page, selector: str, value: str) -> None:
    locator = page.locator(selector)
    await locator.wait_for(state="visible")
    await locator.click()
    await locator.fill(value)
    actual = (await locator.input_value()).strip()
    if actual != value.strip():
        await locator.press("Control+A")
        await locator.type(value)


def _resolve_family_value(family: str, mapping: dict[str, str]) -> Optional[str]:
    if not family:
        return None
    key = _normalise(family)
    if key in mapping:
        return mapping[key]
    # Try exact match without case folding
    for candidate, value in mapping.items():
        if candidate == family.lower() or value.lower() == key:
            return value
    return None


def _extract_panel_text(html: str) -> str:
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(separator=" ", strip=True)
    return re.sub(r"\s+", " ", text)


async def process_row(
    page: Page, row: InputRow, mapping: dict[str, str]
) -> dict[str, str]:
    logging.info(
        "Processing serial=%s product_code=%s product_family=%s",
        row.serial_number or "-",
        row.product_code or "-",
        row.product_family or "-",
    )

    option_value = _resolve_family_value(row.product_family, mapping)
    if option_value is None:
        raise ValueError(
            f"Unknown product family '{row.product_family}' — update RDHC.html or input data"
        )

    await page.select_option(SELECTORS["product_family"], option_value)
    await _fill(page, SELECTORS["product_code"], row.product_code)
    await _fill(page, SELECTORS["serial_number"], row.serial_number)
    await page.click(SELECTORS["submit"])

    try:
        await page.wait_for_timeout(750)
        await page.wait_for_load_state("networkidle")
    except PlaywrightTimeoutError:
        pass

    try:
        panel_html = await page.inner_html(SELECTORS["result_panel"])
    except PlaywrightTimeoutError:
        panel_html = ""

    panel_text = _extract_panel_text(panel_html)
    logging.info("Result for %s: %s", row.serial_number, panel_text or "<no data>")

    return {
        "SerialNumber": row.serial_number,
        "ProductCode": row.product_code,
        "ProductFamily": row.product_family,
        "PanelText": panel_text,
    }


def write_results(path: Path, rows: Iterable[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["SerialNumber", "ProductCode", "ProductFamily", "PanelText"]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


async def main() -> None:
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s"
    )

    family_mapping = load_product_family_map(RDHC_HTML_PATH)
    rows = load_input_rows(AST_INPUT_XLSX)
    logging.info("Loaded %d AST input rows from %s", len(rows), AST_INPUT_XLSX)
    if not rows:
        logging.warning("No AST rows found in %s", AST_INPUT_XLSX)
        return

    results: list[dict[str, str]] = []

    async with async_playwright() as playwright:
        channel: str | None = AST_BROWSER_CHANNEL.strip() or None
        storage_state_path: Path | None
        if AST_STORAGE_STATE.exists():
            storage_state_path = AST_STORAGE_STATE
        else:
            logging.warning(
                "Storage state %s not found. The RDHC session may require a manual login.",
                AST_STORAGE_STATE,
            )
            storage_state_path = None

        try:
            browser, context = await launch_browser(
                playwright,
                headless=AST_HEADLESS,
                channel=channel,
                storage_state_path=storage_state_path,
            )
        except PlaywrightError as exc:
            logging.error(
                "Failed to launch Chromium browser. %s",
                (
                    "Install the Playwright browsers by running 'playwright install chromium' "
                    "or set AST_BROWSER_CHANNEL to a locally installed browser."
                ),
            )
            logging.debug("Playwright launch error: %s", exc)
            return

        try:
            page = await context.new_page()
            await page.goto(AST_PAGE_URL, wait_until="domcontentloaded")

            for row in rows:
                try:
                    result = await process_row(page, row, family_mapping)
                except Exception as exc:  # noqa: BLE001
                    logging.error(
                        "Failed to process serial=%s product=%s: %s",
                        row.serial_number,
                        row.product_code,
                        exc,
                    )
                    results.append(
                        {
                            "SerialNumber": row.serial_number,
                            "ProductCode": row.product_code,
                            "ProductFamily": row.product_family,
                            "PanelText": f"ERROR: {exc}",
                        }
                    )
                    continue

                results.append(result)

        finally:
            await context.close()
            await browser.close()

    write_results(AST_OUTPUT_CSV, results)
    logging.info("Wrote AST toner results to %s", AST_OUTPUT_CSV)


if __name__ == "__main__":
    asyncio.run(main())
