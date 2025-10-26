#!/usr/bin/env python3
"""
WebForms replayer for SingleRequest.aspx (SEARCH only).
Uses Playwright's storage_state.json for cookies (no visible browser).
Reads devices from FIRMWARE_INPUT_XLSX (CSV/XLSX), posts WebForms payloads,
and writes <input>_out.csv with HTTP/status columns.

Env:
  FIRMWARE_INPUT_XLSX=data/firmware_schedule.csv
  FIRMWARE_STORAGE_STATE=storage_state.json

Requires:
  pip install httpx beautifulsoup4 truststore
  (and if using .xlsx) pip install openpyxl
"""

from __future__ import annotations

import csv
import json
import os
import time
from pathlib import Path
from typing import Any, Dict, Iterable, Tuple

# --- Windows / corp TLS trust (safe no-op if unavailable) ---
try:
    import truststore  # type: ignore

    truststore.inject_into_ssl()  # use Windows certificate store
except Exception:
    pass

import httpx
from bs4 import BeautifulSoup, Tag

# ----- Constants from your capture -----
BASE = "https://sgpaphq-epbbcs3.dc01.fujixerox.net"
URL = f"{BASE}/firmware/SingleRequest.aspx"

SEARCH_PANEL = "ctl00$MainContent$searchForm"
SEARCH_TRIGGER = "ctl00$MainContent$btnSearch"

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "X-Requested-With": "XMLHttpRequest",
    "X-MicrosoftAjax": "Delta=true",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    "Origin": BASE,
    "Referer": URL,
}

DEFAULT_OPCO = os.getenv("FIRMWARE_OPCO", "FXAU")
DRY_RUN = os.getenv("FIRMWARE_DRY_RUN", "false").lower() in {"1", "true", "yes"}

INPUT_PATH = Path(os.getenv("FIRMWARE_INPUT_XLSX", "data/firmware_schedule.csv"))
STORAGE_STATE_PATH = Path(os.getenv("FIRMWARE_STORAGE_STATE", "storage_state.json"))


def _as_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ",".join(str(v) for v in value)
    return str(value)


def read_rows(path: Path) -> Iterable[dict]:
    """Yield dicts with serial, product_code, state, opco (CSV or XLSX)."""
    if not path.exists():
        raise FileNotFoundError(f"Input not found: {path}")

    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                yield normalize_row(row)
        return

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:
            raise RuntimeError("openpyxl is required for .xlsx files") from exc

        wb = load_workbook(path, read_only=True)
        try:
            ws = wb.active
            if ws is None:
                raise RuntimeError(f"No active worksheet in workbook: {path}")

            # header row (guard StopIteration)
            header_iter = ws.iter_rows(min_row=1, max_row=1)
            try:
                header_cells = next(header_iter)
            except StopIteration:
                return  # empty sheet

            headers = [
                "" if c.value is None else str(c.value).strip() for c in header_cells
            ]

            for cells in ws.iter_rows(min_row=2, values_only=True):
                row = {
                    (headers[i] or f"col{i}"): ("" if v is None else str(v).strip())
                    for i, v in enumerate(cells)
                }
                yield normalize_row(row)
        finally:
            wb.close()
        return

    raise ValueError(f"Unsupported input type: {path.suffix}")


def normalize_row(raw: dict) -> dict:
    lower = {
        (k or "").strip().lower(): ("" if v is None else str(v).strip())
        for k, v in raw.items()
    }

    def get(*names: str) -> str:
        for n in names:
            if n in lower and lower[n]:
                return lower[n]
        return ""

    return {
        "serial": get("serial", "serialnumber", "serial_number"),
        "product_code": get("product_code", "product", "productcode"),
        "state": get("state", "region"),
        "opco": get("opco", "opcoid", "opco_id") or DEFAULT_OPCO,
    }


# ----- Auth -----
def import_cookies(client: httpx.Client) -> None:
    """Load Playwright storage_state.json cookies into the client."""
    if not STORAGE_STATE_PATH.exists():
        raise FileNotFoundError(f"Missing storage_state.json: {STORAGE_STATE_PATH}")
    payload = json.loads(STORAGE_STATE_PATH.read_text(encoding="utf-8"))
    cookies = payload.get("cookies", [])
    for c in cookies:
        name = c.get("name")
        value = c.get("value")
        domain = c.get("domain") or "sgpaphq-epbbcs3.dc01.fujixerox.net"
        path = c.get("path") or "/"
        if name and value:
            client.cookies.set(name, value, domain=domain, path=path)


# ----- WebForms -----
def extract_hidden(html: str) -> Dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")

    def val(name: str) -> str:
        el = soup.find("input", {"name": name})
        if not el or not isinstance(el, Tag):
            return ""
        return _as_str(el.get("value"))

    fields: Dict[str, str] = {}
    for name in (
        "__VIEWSTATE",
        "__VIEWSTATEGENERATOR",
        "__EVENTVALIDATION",
        "__EVENTTARGET",
        "__EVENTARGUMENT",
        "__LASTFOCUS",
    ):
        fields[name] = val(name)
    return fields


def get_state(client: httpx.Client) -> Dict[str, str]:
    r = client.get(URL, headers={"User-Agent": HEADERS["User-Agent"]})
    if r.status_code == 401:
        raise PermissionError("Unauthorized (401): cookies may be expired or invalid.")
    r.raise_for_status()
    return extract_hidden(r.text)


def post_search(
    client: httpx.Client, opco: str, product_code: str, serial: str
) -> Tuple[int, str]:
    hidden = get_state(client)
    form = {
        "ctl00$ScriptManager1": f"{SEARCH_PANEL}|{SEARCH_TRIGGER}",
        "__EVENTTARGET": hidden.get("__EVENTTARGET", ""),
        "__EVENTARGUMENT": hidden.get("__EVENTARGUMENT", ""),
        "__LASTFOCUS": hidden.get("__LASTFOCUS", ""),
        "__VIEWSTATE": hidden.get("__VIEWSTATE", ""),
        "__VIEWSTATEGENERATOR": hidden.get("__VIEWSTATEGENERATOR", ""),
        "__EVENTVALIDATION": hidden.get("__EVENTVALIDATION", ""),
        "ctl00$MainContent$ddlOpCoID": opco,
        "ctl00$MainContent$ProductCode": product_code,
        "ctl00$MainContent$SerialNumber": serial,
        "ctl00$ucAsync1$hdnTimeout": "30",
        "ctl00$ucAsync1$hdnSetTimeoutID": "4",
        "__ASYNCPOST": "true",
        SEARCH_TRIGGER: "Search",
    }
    if DRY_RUN:
        return 200, "DRY_RUN search"
    r = client.post(URL, data=form, headers=HEADERS, timeout=60.0)
    return r.status_code, r.text


def parse_status(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for sel in [
        "#MainContent_MessageLabel",
        "#MainContent_lblMessage",
        "#MainContent_lblStatus",
    ]:
        node = soup.select_one(sel)
        if node:
            return " ".join(node.get_text(" ", strip=True).split())
    if html.startswith("|"):  # MS AJAX delta format
        upper = html.upper()
        for key in ("SUCCESS", "SCHEDULE", "NOT", "INVALID", "ERROR"):
            if key in upper:
                return key
    return ""


def main() -> None:
    in_path = INPUT_PATH
    out_path = in_path.with_name(in_path.stem + "_out.csv")

    with (
        httpx.Client(follow_redirects=True, timeout=60.0, trust_env=True) as client,
        open(out_path, "w", newline="", encoding="utf-8") as fout,
    ):
        import_cookies(client)

        rows = list(read_rows(in_path))
        if not rows:
            print(f"No rows found in {in_path}")
            return

        fieldnames = [
            "serial",
            "product_code",
            "state",
            "opco",
            "http_status_search",
            "status_text_search",
        ]
        writer = csv.DictWriter(fout, fieldnames=fieldnames)
        writer.writeheader()

        for item in rows:
            serial = item["serial"]
            product = item["product_code"]
            if not serial or not product:
                continue
            code, html = post_search(
                client, item.get("opco") or DEFAULT_OPCO, product, serial
            )
            status_text = parse_status(html)
            writer.writerow(
                {
                    "serial": serial,
                    "product_code": product,
                    "state": item.get("state", ""),
                    "opco": item.get("opco", ""),
                    "http_status_search": code,
                    "status_text_search": status_text,
                }
            )
            time.sleep(0.2)

    print(f"✅ Done. Wrote: {out_path}")


if __name__ == "__main__":
    main()
