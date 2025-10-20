"""Automate toner availability lookups and capture the results in Excel.

The script drives the Fuji Xerox parts status page via Playwright, reads input
rows from ``input.xlsx`` and persists the enriched results to ``output.xlsx``.
Robust error handling is included so transient site issues or unexpected input
structures do not crash the run.
"""

import asyncio
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple, cast

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import Page, async_playwright

# Input and output file paths
DOWNLOAD_DIR = Path("downloads")
OUTPUT_XLSX = Path("output.xlsx")  # Where results will be saved
SCREENSHOTS_DIR = Path("screenshots")

# Selectors for the input fields and search button
SELECTORS = {
    "product_family": "#MainContent_txtProductFamily",
    "product_code": "#MainContent_txtProductCode",
    "serial_number": "#MainContent_txtSerialNumber",
    "search_btn": "#MainContent_btnSearch",
    "results_table": "#MainContent_gvResults",
}

PAGE_URL = "https://sgpaphq-epbbcs3.dc01.fujixerox.net/rdhc/PartStatuses.aspx"
STORAGE_STATE_ENV = "AST_TONER_STORAGE_STATE"


logger = logging.getLogger(__name__)


def _active_sheet(wb: Workbook) -> Worksheet:
    """Return the active worksheet, raising if the workbook is empty."""

    ws = wb.active
    if ws is None:
        raise ValueError("Workbook has no active worksheet")
    return cast(Worksheet, ws)


def _resolve_input_workbook() -> Path:
    """Locate the workbook to ingest, preferring an explicit override."""

    env_path = os.environ.get("AST_TONER_INPUT")
    if env_path:
        candidate = Path(env_path).expanduser()
        if candidate.is_file():
            return candidate
        raise FileNotFoundError(
            f"AST_TONER_INPUT was set to '{candidate}', but the file does not exist."
        )

    if not DOWNLOAD_DIR.exists():
        raise FileNotFoundError(f"Download directory '{DOWNLOAD_DIR}' does not exist.")

    candidates = sorted(
        (
            path
            for path in DOWNLOAD_DIR.glob("*.xlsx")
            if path.is_file() and not path.name.startswith("~")
        ),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    if not candidates:
        raise FileNotFoundError(f"No .xlsx files found in '{DOWNLOAD_DIR}'.")

    return candidates[0]


def _resolve_storage_state() -> Optional[Path]:
    """Return a valid storage state file when available.

    When ``AST_TONER_STORAGE_STATE`` is set we require that file to exist.
    Otherwise we fall back to ``storage_state.json`` next to the script if it is
    present. Returning ``None`` signals that no storage state should be passed
    to Playwright.
    """

    env_path = os.environ.get(STORAGE_STATE_ENV)
    if env_path:
        candidate = Path(env_path).expanduser()
        if candidate.is_file():
            return candidate
        raise FileNotFoundError(
            f"{STORAGE_STATE_ENV} was set to '{candidate}', but the file does not exist."
        )

    default_path = Path("storage_state.json")
    if default_path.is_file():
        return default_path

    logger.warning(
        "No Playwright storage state found. Proceeding without persisted login; a fresh login will be required."
    )
    return None


def parse_html_table(html: str) -> Tuple[List[str], List[List[str]]]:
    """Extract header and body rows from an HTML table snippet."""

    soup = BeautifulSoup(html, "html.parser")
    table = soup.select_one("table")
    if table is None:
        raise ValueError("No <table> found in the HTML!")

    headers = [th.get_text(strip=True) for th in table.select("tr th")]
    rows: List[List[str]] = []
    for tr in table.select("tr")[1:]:
        cells = [td.get_text(strip=True) for td in tr.select("td")]
        if cells:
            rows.append(cells)

    return headers, rows


def _cell_value(row: Sequence, index: int) -> str:
    """Safely read and normalise a cell value from ``row``."""

    if index >= len(row):
        return ""
    value = row[index].value
    return str(value).strip() if value is not None else ""


def _row_has_data(values: Iterable[str]) -> bool:
    """Return True when at least one value contains user-provided data."""

    return any(value for value in values)


def _ensure_screenshot_dir() -> None:
    """Ensure the screenshot directory exists prior to screenshot capture."""

    SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)


def _screenshot_path(row_index: int, stage: str) -> Path:
    """Build a timestamped screenshot path for a given row and stage."""

    timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S-%f")
    return SCREENSHOTS_DIR / f"row{row_index:04d}-{stage}-{timestamp}.png"


async def _fill_input(page: Page, selector: str, value: str, field_name: str) -> bool:
    """Populate an input and confirm the DOM reflects the expected data."""

    locator = page.locator(selector)
    await locator.wait_for(state="visible")
    await locator.click()

    target_value = value or ""
    await locator.fill(target_value)

    actual_value = (await locator.input_value()).strip()
    if target_value and not actual_value:
        logger.warning(
            "%s field appeared blank after initial fill; retrying with type().",
            field_name,
        )
        await locator.fill("")
        await locator.type(target_value)
        actual_value = (await locator.input_value()).strip()

    if target_value and not actual_value:
        logger.error("Failed to set %s field to '%s'", field_name, target_value)
        return False

    await locator.press("Tab")
    return True


async def main():
    logging.basicConfig(
        format="%(asctime)s %(levelname)s %(message)s",
        level=logging.INFO,
    )

    try:
        input_xlsx = _resolve_input_workbook()
    except FileNotFoundError as exc:
        logger.error("Could not locate input workbook: %s", exc)
        raise

    logger.info("Loading input workbook: %s", input_xlsx)
    wb = load_workbook(input_xlsx)
    ws: Worksheet = _active_sheet(wb)

    out_wb = Workbook()
    out_ws: Worksheet = _active_sheet(out_wb)
    out_ws.title = "Results"

    _ensure_screenshot_dir()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        storage_state_path = _resolve_storage_state()
        context = (
            await browser.new_context(storage_state=str(storage_state_path))
            if storage_state_path
            else await browser.new_context()
        )
        page = await context.new_page()
        await page.goto(PAGE_URL)

        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            serial_number = _cell_value(row, 0)
            product_code = _cell_value(row, 1)
            product_family = _cell_value(row, 6)

            if not _row_has_data((serial_number, product_code, product_family)):
                logger.info("Skipping empty row %s", idx)
                continue

            logger.info(
                "Processing row %s: family=%s, code=%s, serial=%s",
                idx,
                product_family or "-",
                product_code or "-",
                serial_number or "-",
            )

            family_ok = await _fill_input(
                page, SELECTORS["product_family"], product_family, "Product Family"
            )
            code_ok = await _fill_input(
                page, SELECTORS["product_code"], product_code, "Product Code"
            )
            serial_ok = await _fill_input(
                page, SELECTORS["serial_number"], serial_number, "Serial Number"
            )

            if family_ok and code_ok and serial_ok:
                filled_path = _screenshot_path(idx, "filled")
                await page.screenshot(path=str(filled_path), full_page=True)
                logger.info("Captured filled-fields screenshot: %s", filled_path)
            else:
                logger.warning(
                    "Skipping pre-search screenshot for row %s due to input issues", idx
                )

            await page.click(SELECTORS["search_btn"])

            try:
                await page.wait_for_selector(SELECTORS["results_table"], timeout=10000)
                table_html = await page.inner_html(SELECTORS["results_table"])
                headers, data_rows = parse_html_table(table_html)
                results_path = _screenshot_path(idx, "results")
                await page.screenshot(path=str(results_path), full_page=True)
                logger.info("Captured results screenshot: %s", results_path)
            except PlaywrightTimeoutError:
                logger.warning("Timed out waiting for results table on row %s", idx)
                headers, data_rows = ["Status"], [["Lookup timed out"]]
                timeout_path = _screenshot_path(idx, "timeout")
                await page.screenshot(path=str(timeout_path), full_page=True)
                logger.info("Captured timeout screenshot for row %s: %s", idx, timeout_path)
            except ValueError as exc:
                logger.warning("Failed to parse table for row %s: %s", idx, exc)
                headers, data_rows = ["Status"], [["No results table found"]]
                missing_path = _screenshot_path(idx, "no-table")
                await page.screenshot(path=str(missing_path), full_page=True)
                logger.info(
                    "Captured missing-table screenshot for row %s: %s", idx, missing_path
                )

            if idx == 2:
                out_ws.append(
                    ["Input Row", "Product Family", "Product Code", "Serial Number"]
                    + headers
                )

            if not data_rows:
                data_rows = [["No results returned"]]

            for data_row in data_rows:
                out_ws.append(
                    [idx, product_family, product_code, serial_number] + data_row
                )

        logger.info("Saving results workbook: %s", OUTPUT_XLSX)
        out_wb.save(OUTPUT_XLSX)
        await browser.close()


# To run in a notebook or interactive environment, use:
# asyncio.get_event_loop().run_until_complete(main())
# Otherwise, use:
if __name__ == "__main__":
    asyncio.run(main())
