#!/usr/bin/env python3
"""
Headless Playwright replayer for SingleRequest.aspx (SEARCH + SCHEDULE).

- Uses the browser (Chromium/Edge) so NTLM/Negotiate and TLS Just Work™.
- Loads cookies from storage_state.json (optional but recommended).
- Reads devices from FIRMWARE_INPUT_XLSX (CSV/XLSX).
- SEARCH: posts the UpdatePanel payload via in-page fetch() (browser cookies/auth).
- SCHEDULE: performs a real WebForms postback (robust click with manual form fallback).
- Writes <input>_out.csv with search & schedule results.

Env:
  FIRMWARE_INPUT_XLSX=data/firmware_schedule.csv
  FIRMWARE_STORAGE_STATE=storage_state.json
  FIRMWARE_BROWSER_CHANNEL=msedge
  FIRMWARE_AUTH_ALLOWLIST=*.fujixerox.net,*.xerox.com
  FIRMWARE_OPCO=FXAU
  FIRMWARE_HEADLESS=true
  # Optional:
  FIRMWARE_TIME_VALUE=03
  FIRMWARE_DAYS_MIN=3
  FIRMWARE_DAYS_MAX=6
  FIRMWARE_DEBUG_TZ=0   # set to 1 to print timezone options/selection
"""

from __future__ import annotations

import asyncio
import csv
import os
import random
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, Tuple, List

from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, Error as PWError  # type: ignore

# ---------- Constants ----------
BASE = "https://sgpaphq-epbbcs3.dc01.fujixerox.net"
URL = f"{BASE}/firmware/SingleRequest.aspx"

# UpdatePanel targets (from captured posts)
SEARCH_PANEL = "ctl00$MainContent$searchForm"
SEARCH_TRIGGER = "ctl00$MainContent$btnSearch"
SCHEDULE_PANEL = "ctl00$MainContent$pnlFWTimesssss"
SCHEDULE_TRIGGER = "ctl00$MainContent$submitButton"

# Headers for XHR; Origin/Referer/Cookies are handled by the browser.
XHR_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "X-Requested-With": "XMLHttpRequest",
    "X-MicrosoftAjax": "Delta=true",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
}

DEFAULT_OPCO = os.getenv("FIRMWARE_OPCO", "FXAU")
INPUT_PATH = Path(os.getenv("FIRMWARE_INPUT_XLSX", "data/firmware_schedule.csv"))
STORAGE_STATE_PATH = Path(os.getenv("FIRMWARE_STORAGE_STATE", "storage_state.json"))
BROWSER_CHANNEL = os.getenv("FIRMWARE_BROWSER_CHANNEL", "msedge") or None
ALLOWLIST = os.getenv("FIRMWARE_AUTH_ALLOWLIST", "*.fujixerox.net,*.xerox.com")
HEADLESS = os.getenv("FIRMWARE_HEADLESS", "true").lower() in {"1", "true", "yes"}
DEBUG_TZ = os.getenv("FIRMWARE_DEBUG_TZ", "0").lower() in {"1", "true", "yes"}

PREFERRED_TIME_VALUE = (os.getenv("FIRMWARE_TIME_VALUE", "03") or "03").strip()
DAYS_MIN = int(os.getenv("FIRMWARE_DAYS_MIN", "3"))
DAYS_MAX = int(os.getenv("FIRMWARE_DAYS_MAX", "6"))

# AU state → timezone dropdown value (server expects the <option value>, not label)
STATE_TZ = {
    "ACT": "+11:00",
    "NSW": "+11:00",
    "VIC": "+11:00",
    "TAS": "+11:00",
    "QLD": "+10:00",
    "SA": "+10:30",
    "NT": "+09:30",
}


# ---------- Input helpers ----------
def normalize_row(raw: dict) -> dict:
    lower = {
        (k or "").strip().lower(): "" if v is None else str(v).strip()
        for k, v in raw.items()
    }

    def get(*names: str) -> str:
        for n in names:
            if n in lower and lower[n]:
                return lower[n]
        return ""

    return {
        "serial": get("serial", "serialnumber", "serial_number"),
        "product_code": get("product_code", "product", "productcode"),
        "state": get("state", "region").upper(),
        "opco": get("opco", "opcoid", "opco_id") or DEFAULT_OPCO,
    }


def read_rows(path: Path) -> Iterable[dict]:
    if not path.exists():
        raise FileNotFoundError(f"Input not found: {path}")

    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                yield normalize_row(row)
        return

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:
            raise RuntimeError("openpyxl is required for .xlsx files") from exc

        wb = load_workbook(path, read_only=True)
        try:
            ws = wb.active
            if ws is None:
                raise RuntimeError(f"No active worksheet in workbook: {path}")
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            headers = [
                "" if c.value is None else str(c.value).strip() for c in header_cells
            ]
            for cells in ws.iter_rows(min_row=2, values_only=True):
                row = {
                    (headers[i] or f"col{i}"): ("" if v is None else str(v).strip())
                    for i, v in enumerate(cells)
                }
                yield normalize_row(row)
        finally:
            wb.close()
        return

    raise ValueError(f"Unsupported input type: {path.suffix}")


# ---------- MicrosoftAjax delta parsing ----------
def _extract_from_msajax_delta(delta: str) -> str:
    if not delta.startswith("|"):
        return ""
    tokens: List[str] = delta.split("|")
    i = 0
    snippets: List[str] = []
    while i < len(tokens):
        if tokens[i] == "updatePanel" and i + 3 < len(tokens):
            html = tokens[i + 3]
            snippets.append(html)
            i += 4
            continue
        i += 1
    for html in snippets:
        soup = BeautifulSoup(html, "html.parser")
        for sel in [
            "#MainContent_MessageLabel",
            "#MainContent_lblMessage",
            "#MainContent_lblStatus",
        ]:
            node = soup.select_one(sel)
            if node:
                txt = " ".join(node.get_text(" ", strip=True).split())
                if txt:
                    return txt
    for html in snippets:
        cleaned = " ".join(
            BeautifulSoup(html, "html.parser").get_text(" ", strip=True).split()
        )
        if cleaned:
            return cleaned[:300]
    return ""


def parse_status_from_html(html: str) -> str:
    if html.startswith("|"):
        return _extract_from_msajax_delta(html)
    soup = BeautifulSoup(html, "html.parser")
    for sel in [
        "#MainContent_MessageLabel",
        "#MainContent_lblMessage",
        "#MainContent_lblStatus",
    ]:
        node = soup.select_one(sel)
        if node:
            return " ".join(node.get_text(" ", strip=True).split())
    upper = html.upper()
    for key in ("SUCCESS", "SCHEDULE", "NOT", "INVALID", "ERROR"):
        if key in upper:
            return key
    return ""


# ---------- Scheduling helpers ----------
def pick_schedule_date() -> str:
    start = datetime.now().date() + timedelta(days=DAYS_MIN)
    end = datetime.now().date() + timedelta(days=DAYS_MAX)
    span = (end - start).days
    offset = random.randint(0, max(0, span))
    return (start + timedelta(days=offset)).strftime("%Y-%m-%d")


def timezone_for_state(state: str) -> str:
    return STATE_TZ.get(state.upper(), "+11:00")


async def select_timezone_on_page(
    page, desired_value: str, label_hint: str | None = None
) -> str:
    """
    Selects timezone in the DOM so VIEWSTATE reflects the selection.
    Returns the actually selected option value.
    """
    # small wait + force-enable
    await page.wait_for_timeout(100)
    await page.evaluate("""
      () => {
        const sel = document.querySelector('select#MainContent_ddlTimeZone');
        if (sel) { sel.removeAttribute('disabled'); sel.disabled = false; }
      }
    """)

    ok = await page.evaluate(
        """
        (val) => {
          const sel = document.querySelector('select#MainContent_ddlTimeZone');
          if (!sel) return false;
          sel.value = val;
          if (sel.value === val) {
            sel.dispatchEvent(new Event('change', { bubbles: true }));
            return true;
          }
          return false;
        }
    """,
        desired_value,
    )
    if ok:
        return (
            await page.eval_on_selector(
                "select#MainContent_ddlTimeZone", "el => el.value"
            )
            or desired_value
        )

    if label_hint:
        matched_val = await page.evaluate(
            """
            (needle) => {
              const sel = document.querySelector('select#MainContent_ddlTimeZone');
              if (!sel) return '';
              const m = Array.from(sel.options).find(o => (o.text||'').toUpperCase().includes(needle.toUpperCase()));
              return m ? m.value : '';
            }
        """,
            label_hint,
        )
        if matched_val:
            ok2 = await page.evaluate(
                """
                (val) => {
                  const sel = document.querySelector('select#MainContent_ddlTimeZone');
                  if (!sel) return false;
                  sel.value = val;
                  if (sel.value === val) {
                    sel.dispatchEvent(new Event('change', { bubbles: true }));
                    return true;
                  }
                  return false;
                }
            """,
                matched_val,
            )
            if ok2:
                return matched_val

    fallback = await page.evaluate("""
        () => {
          const sel = document.querySelector('select#MainContent_ddlTimeZone');
          if (!sel) return '';
          const opt = Array.from(sel.options).find(o => o.value);
          if (!opt) return '';
          sel.value = opt.value;
          sel.dispatchEvent(new Event('change', { bubbles: true }));
          return sel.value;
        }
    """)
    return fallback or desired_value


async def debug_dump_timezone(page) -> None:
    if not DEBUG_TZ:
        return
    data = await page.evaluate("""
      () => {
        const sel = document.querySelector('#MainContent_ddlTimeZone');
        if (!sel) return {selected: null, options: []};
        return {
          selected: { value: sel.value, text: sel.options[sel.selectedIndex]?.text || '' },
          options: Array.from(sel.options).map(o => ({ value: o.value, text: o.text }))
        };
      }
    """)
    print("TZ selected:", data.get("selected"))
    for o in data.get("options", []):
        print("TZ option:", o)


# ---------- WebForms via Playwright ----------
async def get_state(page) -> Dict[str, str]:
    # Navigating ensures IWA handshake + the form exists
    await page.goto(URL, wait_until="domcontentloaded")
    names = [
        "__VIEWSTATE",
        "__VIEWSTATEGENERATOR",
        "__EVENTVALIDATION",
        "__EVENTTARGET",
        "__EVENTARGUMENT",
        "__LASTFOCUS",
    ]
    values: Dict[str, str] = {}
    for name in names:
        try:
            val = await page.eval_on_selector(f'input[name="{name}"]', "el => el.value")
        except PWError:
            val = ""
        values[name] = val or ""
    return values


def urlencode_form(d: Dict[str, str]) -> str:
    from urllib.parse import urlencode

    return urlencode(d)


async def post_search(
    page, hidden: Dict[str, str], opco: str, product_code: str, serial: str
) -> Tuple[int, str]:
    form = {
        "ctl00$ScriptManager1": f"{SEARCH_PANEL}|{SEARCH_TRIGGER}",
        "__EVENTTARGET": hidden.get("__EVENTTARGET", ""),
        "__EVENTARGUMENT": hidden.get("__EVENTARGUMENT", ""),
        "__LASTFOCUS": hidden.get("__LASTFOCUS", ""),
        "__VIEWSTATE": hidden.get("__VIEWSTATE", ""),
        "__VIEWSTATEGENERATOR": hidden.get("__VIEWSTATEGENERATOR", ""),
        "__EVENTVALIDATION": hidden.get("__EVENTVALIDATION", ""),
        "ctl00$MainContent$ddlOpCoID": opco,
        "ctl00$MainContent$ProductCode": product_code,
        "ctl00$MainContent$SerialNumber": serial,
        "ctl00$ucAsync1$hdnTimeout": "30",
        "ctl00$ucAsync1$hdnSetTimeoutID": "4",
        "__ASYNCPOST": "true",
        SEARCH_TRIGGER: "Search",
    }
    payload = urlencode_form(form)
    result = await page.evaluate(
        """async ({url, headers, body}) => {
            const res = await fetch(url, { method: 'POST', headers, body, credentials: 'include' });
            const text = await res.text();
            return { status: res.status, text };
        }""",
        {"url": URL, "headers": XHR_HEADERS, "body": payload},
    )
    return int(result["status"]), str(result["text"])


async def wait_for_schedule_controls_enabled(page) -> None:
    # Wait until all three are present and enabled (after SEARCH UpdatePanel refresh)
    await page.wait_for_selector("#MainContent_txtDateTime:enabled", timeout=15000)
    await page.wait_for_selector("#MainContent_ddlScheduleTime:enabled", timeout=15000)
    await page.wait_for_selector("#MainContent_ddlTimeZone:enabled", timeout=15000)


async def force_enable_controls(page) -> None:
    # In case the page kept them disabled, force-enable so they submit
    await page.evaluate("""
      () => {
        const drop = el => { if (!el) return; el.removeAttribute('disabled'); el.disabled = false; el.readOnly = false; };
        drop(document.querySelector('#MainContent_txtDateTime'));
        drop(document.querySelector('#MainContent_ddlScheduleTime'));
        drop(document.querySelector('#MainContent_ddlTimeZone'));
      }
    """)


async def dom_submit_schedule(
    page,
    opco: str,
    product_code: str,
    serial: str,
    date_iso: str,
    time_val: str,
    tz_val: str,
) -> tuple[int, str]:
    """
    Fill fields in the existing DOM (produced by SEARCH) and trigger a real WebForms postback.
    Handles both UpdatePanel (no navigation) and full postback (navigation).
    Returns (status_code, parsed_status_text).
    """
    # DO NOT navigate here—stay on the current DOM unlocked by SEARCH.
    await force_enable_controls(page)

    # Fill inputs/selects; trigger change so WebForms tracks state
    await page.evaluate(
        """
        ({opco, product, serial, date, time, tz}) => {
          const setVal = (sel, val) => {
            const el = document.querySelector(sel);
            if (!el) return;
            el.value = val;
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
          };
          setVal('select#MainContent_ddlOpCoID', opco);
          setVal('#MainContent_ProductCode', product);
          setVal('#MainContent_SerialNumber', serial);
          setVal('#MainContent_txtDateTime', date);
          setVal('select#MainContent_ddlScheduleTime', time);
          setVal('select#MainContent_ddlTimeZone', tz);
        }
    """,
        {
            "opco": opco,
            "product": product_code,
            "serial": serial,
            "date": date_iso,
            "time": time_val,
            "tz": tz_val,
        },
    )

    # Verify values stuck; if date looks yyyy-mm-dd and page expects dd/MM/yyyy, rewrite it
    ok_set = await page.evaluate("""
      () => {
        const d = document.querySelector('#MainContent_txtDateTime')?.value || '';
        const t = document.querySelector('select#MainContent_ddlScheduleTime')?.value || '';
        const z = document.querySelector('select#MainContent_ddlTimeZone')?.value || '';
        return { d, t, z };
      }
    """)
    if ok_set and isinstance(ok_set, dict):
        d = ok_set.get("d") or ""
        if d and "-" in d and len(d.split("-")) == 3:
            y, m, dd = d.split("-")
            ddmmyyyy = f"{dd}/{m}/{y}"
            await page.evaluate(
                """
              (val) => {
                const el = document.querySelector('#MainContent_txtDateTime');
                if (!el) return;
                el.value = val;
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
              }
            """,
                ddmmyyyy,
            )

    await page.wait_for_timeout(150)  # let client-side validators settle

    if DEBUG_TZ:
        await debug_dump_timezone(page)

    # Try to click schedule (async delta or full postback)
    selectors = [
        'input[name="ctl00$MainContent$submitButton"]',
        "#MainContent_submitButton",
        'input[type="submit"][value="Schedule"]',
        'input[value="Schedule"]',
        "button#MainContent_submitButton",
        'button[name="ctl00$MainContent$submitButton"]',
        '//input[contains(@value,"Schedule")]',
        '//button[contains(normalize-space(),"Schedule")]',
    ]
    clicked = False
    for sel in selectors:
        try:
            loc = page.locator(sel)
            if await loc.count() == 0:
                continue
            await loc.first.wait_for(state="visible", timeout=3000)
            if await loc.first.is_disabled():
                continue
            await loc.first.scroll_into_view_if_needed()
            # Race: maybe nav, maybe just partial update
            nav_wait = page.wait_for_load_state("domcontentloaded", timeout=15000)
            msg_wait = page.wait_for_selector(
                "#MainContent_MessageLabel, #MainContent_lblMessage, #MainContent_lblStatus",
                timeout=15000,
            )
            await loc.first.click(timeout=3000)
            try:
                await asyncio.wait(
                    [asyncio.create_task(nav_wait), asyncio.create_task(msg_wait)],
                    return_when=asyncio.FIRST_COMPLETED,
                    timeout=15000,
                )
            except Exception:
                pass
            clicked = True
            break
        except Exception:
            continue

    if not clicked:
        # Manual full postback without Sys.WebForms (avoids strict-mode issues)
        await page.evaluate(
            """
            (eventTarget) => {
              var f = document.forms && document.forms[0];
              if (!f) return;
              var et = f.__EVENTTARGET || f.querySelector('input[name="__EVENTTARGET"]');
              if (!et) { et = document.createElement('input'); et.type='hidden'; et.name='__EVENTTARGET'; f.appendChild(et); }
              et.value = eventTarget;
              var ea = f.__EVENTARGUMENT || f.querySelector('input[name="__EVENTARGUMENT"]');
              if (!ea) { ea = document.createElement('input'); ea.type='hidden'; ea.name='__EVENTARGUMENT'; f.appendChild(ea); }
              ea.value = '';
              f.submit();
            }
        """,
            "ctl00$MainContent$submitButton",
        )
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception:
            pass

    # Helper: safe status read that tolerates navigation during the loop
    async def read_status_safe() -> str:
        try:
            for sel in [
                "#MainContent_MessageLabel",
                "#MainContent_lblMessage",
                "#MainContent_lblStatus",
            ]:
                el = await page.query_selector(sel)
                if el:
                    txt = (await el.inner_text()).strip()
                    if txt:
                        return " ".join(txt.split())
        except PWError:
            try:
                await page.wait_for_load_state("domcontentloaded", timeout=5000)
            except Exception:
                pass
        return ""

    # Poll for message (works for both partial update and full reload)
    for _ in range(60):  # ~12s
        msg = await read_status_safe()
        if msg:
            return 200, msg
        await page.wait_for_timeout(200)

    return 200, ""


# ---------- Main ----------
async def main() -> None:
    out_path = INPUT_PATH.with_name(INPUT_PATH.stem + "_out.csv")
    browser_args = [
        f"--auth-server-allowlist={ALLOWLIST}",
        f"--auth-negotiate-delegate-allowlist={ALLOWLIST}",
    ]

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=HEADLESS, channel=BROWSER_CHANNEL, args=browser_args
        )
        context_kwargs: Dict[str, Any] = {}
        if STORAGE_STATE_PATH.exists():
            context_kwargs["storage_state"] = str(STORAGE_STATE_PATH)
        context = await browser.new_context(**context_kwargs)
        page = await context.new_page()
        page.set_default_navigation_timeout(45_000)
        page.set_default_timeout(45_000)

        rows = list(read_rows(INPUT_PATH))
        if not rows:
            print(f"No rows found in {INPUT_PATH}")
            await context.close()
            await browser.close()
            return

        with out_path.open("w", newline="", encoding="utf-8") as fout:
            fieldnames = [
                "serial",
                "product_code",
                "state",
                "opco",
                "http_status_search",
                "status_text_search",
                "http_status_schedule",
                "status_text_schedule",
                "scheduled_date",
                "scheduled_time",
                "timezone_value",
            ]
            writer = csv.DictWriter(fout, fieldnames=fieldnames)
            writer.writeheader()

            for item in rows:
                serial = item["serial"]
                product = item["product_code"]
                if not serial or not product:
                    continue

                # 1) SEARCH via UpdatePanel XHR (this navigation inside get_state is OK)
                hidden = await get_state(page)
                code_s, html_s = await post_search(
                    page, hidden, item.get("opco") or DEFAULT_OPCO, product, serial
                )
                status_s = parse_status_from_html(html_s)

                # 2) Wait for schedule controls to be enabled in THIS DOM (no navigation!)
                await wait_for_schedule_controls_enabled(page)

                # 3) Prepare schedule inputs/values
                date_iso = pick_schedule_date()
                time_val = PREFERRED_TIME_VALUE
                desired_tz_val = timezone_for_state(item.get("state", ""))

                # Ensure timezone is actually selected in this DOM (so VIEWSTATE agrees)
                label_hint = (
                    "Canberra, Melbourne, Sydney"
                    if desired_tz_val == "+11:00"
                    else None
                )
                actual_tz_val = await select_timezone_on_page(
                    page, desired_tz_val, label_hint
                )

                # 4) SCHEDULE via real postback on the same DOM
                code_c, status_c = await dom_submit_schedule(
                    page,
                    item.get("opco") or DEFAULT_OPCO,
                    product,
                    serial,
                    date_iso,
                    time_val,
                    actual_tz_val,
                )

                writer.writerow(
                    {
                        "serial": serial,
                        "product_code": product,
                        "state": item.get("state", ""),
                        "opco": item.get("opco", ""),
                        "http_status_search": code_s,
                        "status_text_search": status_s,
                        "http_status_schedule": code_c,
                        "status_text_schedule": status_c,
                        "scheduled_date": date_iso,
                        "scheduled_time": time_val,
                        "timezone_value": actual_tz_val,
                    }
                )

        await context.close()
        await browser.close()

    print(f"Done. Wrote: {out_path}")


if __name__ == "__main__":
    asyncio.run(main())
